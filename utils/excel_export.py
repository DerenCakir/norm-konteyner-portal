"""Excel export helpers for the admin panel.

Per-week workbook (``build_week_excel``) layout:

    1. Renk Kırılımı           — selected week, per (dept × color) detail
    2. Üretim Yeri Kırılımı    — selected week, per-department aggregate
    3. Üretim Yeri Özeti       — selected week, per-site aggregate with %
                                 and ton/dolu KPI
    4. Renk Özeti              — selected week, per-color aggregate
    5. Grafikler               — ALL weeks, charts only
    6. Üretim Yeri Karşılaştırma — ALL weeks, site-summary tables placed
                                 side by side for visual comparison

Long-format export (``build_all_weeks_excel``) is a separate workbook
that lists every (week × site × dept × color) row in one sheet for
PivotTable consumption — unchanged by this refactor.
"""

from __future__ import annotations

from datetime import datetime, date, timedelta
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import NumFmt
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
    RichTextProperties,
)
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from utils.week import now_tr


_STATUS_LABEL = {
    "submitted": "Zamanında",
    "late_submitted": "Geç giriş",
    "draft": "Taslak",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F3A8A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL = PatternFill("solid", fgColor="3B5BB1")
_ZEBRA_FILL = PatternFill("solid", fgColor="F1F5F9")
_LATE_FILL = PatternFill("solid", fgColor="FEF3C7")
_TOTAL_FILL = PatternFill("solid", fgColor="E0EAF8")
_TOTAL_FONT = Font(bold=True, color="1F3A8A", size=11)
_THIN = Side(style="thin", color="D8E4F2")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

# Brand-ish hex codes for color-coded chart series. Anything outside this
# set falls back to a neutral gray so the chart still renders.
_COLOR_HEX = {
    "Mavi": "2563EB",
    "Turuncu": "F97316",
    "Yeşil": "22C55E",
    "Gri": "6B7280",
    "MS Vida": "854D0E",
    "Sarı": "EAB308",
    "Mor": "8B5CF6",
    "Kırmızı": "DC2626",
    "Siyah": "111827",
}


# ---------------------------------------------------------------------------
# Small helpers (kept stable across refactor)
# ---------------------------------------------------------------------------

def _fmt_ts(value: Any) -> str:
    if value in (None, ""):
        return ""
    s = str(value)
    if "T" in s:
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return dt.strftime("%Y-%m-%d %H:%M")
        except ValueError:
            return s[:16].replace("T", " ")
    return s


def _autofit(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 42)


_TR_MONTHS_SHORT = [
    "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]


def _week_iso_to_human(week_iso: str) -> tuple[str, str, int]:
    """Return ``(hafta_aralığı, ay, yıl)`` strings for the given ISO week.

    Falls back to the raw code if parsing fails.
    """
    try:
        year_str, week_str = week_iso.split("-W")
        year = int(year_str)
        week = int(week_str)
        monday = date.fromisocalendar(year, week, 1)
        sunday = monday + timedelta(days=6)
        if monday.year == sunday.year and monday.month == sunday.month:
            label = f"{monday.day:02d}-{sunday.day:02d} {_TR_MONTHS_SHORT[monday.month - 1]} {monday.year}"
        elif monday.year == sunday.year:
            label = (
                f"{monday.day:02d} {_TR_MONTHS_SHORT[monday.month - 1]} - "
                f"{sunday.day:02d} {_TR_MONTHS_SHORT[sunday.month - 1]} {monday.year}"
            )
        else:
            label = (
                f"{monday.day:02d} {_TR_MONTHS_SHORT[monday.month - 1]} {monday.year} - "
                f"{sunday.day:02d} {_TR_MONTHS_SHORT[sunday.month - 1]} {sunday.year}"
            )
        return label, _TR_MONTHS_SHORT[monday.month - 1], monday.year
    except Exception:
        return week_iso, "", 0


def _style_header_row(
    ws, header_count: int, wrap_text: bool = False, row_height: int = 26,
) -> None:
    """Apply standard header styling to the first row of `ws`.

    ``wrap_text=True`` durumunda başlık hücreleri metni kaydır
    olarak hizalanır; ``row_height`` da yazıyı barındıracak şekilde
    büyütülmeli (default 26 tek satır için yeterli, çift satıra
    kıracak başlıklar için 40+ önerilir).
    """
    align = (
        Alignment(horizontal="center", vertical="center", wrap_text=True)
        if wrap_text else _CENTER
    )
    for col_idx in range(1, header_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = align
        cell.border = _BORDER
    ws.row_dimensions[1].height = row_height
    ws.freeze_panes = "A2"


def _set_bar_series_color(series, hex_code: str) -> None:
    """Color a bar/column series solidly. openpyxl needs this dance."""
    series.graphicalProperties = GraphicalProperties(solidFill=hex_code)


def _set_line_series_color(series, hex_code: str) -> None:
    """Color a line series. Solid fill applies to markers; line color
    needs its own LineProperties."""
    gp = GraphicalProperties()
    gp.line = LineProperties(solidFill=hex_code)
    series.graphicalProperties = gp


# ---------------------------------------------------------------------------
# Per-week sheet builders
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Dashboard — one-page summary view (KPI cards + mini chart + top sites)
# ---------------------------------------------------------------------------

def _compute_week_kpis(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Aggregate the per-(site×dept×color) rows down to the KPI numbers
    surfaced on the Dashboard (matches the portal Analiz page logic).
    """
    total_empty = sum(int(r.get("Boş") or 0) for r in rows)
    total_wip = sum(int(r.get("Proseste") or 0) for r in rows)
    total_full = sum(int(r.get("Dolu") or 0) for r in rows)
    total_kanban = sum(int(r.get("Kanban") or 0) for r in rows)
    total_scrap = sum(int(r.get("Hurda") or 0) for r in rows)
    # Yeni Toplam Konteyner tanımı: Boş + WIP + Dolu + Hurda.
    total_containers = total_empty + total_wip + total_full + total_scrap

    # Tonnage: once per submission_id (color rows share the submission tonnage).
    seen_subs: set[int] = set()
    total_tonnage = 0.0
    site_tonnage: dict[str, float] = {}
    for r in rows:
        sub_id = r.get("Submission ID")
        if sub_id is None or sub_id in seen_subs:
            continue
        seen_subs.add(sub_id)
        ton = r.get("Gerçekleşen Tonaj")
        if ton is None:
            continue
        try:
            ton_f = float(ton)
        except (TypeError, ValueError):
            continue
        total_tonnage += ton_f
        site = r.get("Üretim Yeri") or ""
        site_tonnage[site] = site_tonnage.get(site, 0.0) + ton_f

    # Per-site full container counts (across all colors).
    site_full: dict[str, int] = {}
    for r in rows:
        site = r.get("Üretim Yeri") or ""
        site_full[site] = site_full.get(site, 0) + int(r.get("Dolu") or 0)

    # Ortalama Dolu Konteyner Ağırlığı: average of per-site (ton/full) ratios.
    site_ratios: list[float] = []
    for site, full in site_full.items():
        if full <= 0:
            continue
        ton = site_tonnage.get(site, 0.0)
        site_ratios.append(ton * 1000.0 / full)
    avg_kg_per_full = sum(site_ratios) / len(site_ratios) if site_ratios else 0.0

    kanban_pct = (total_kanban / total_full * 100.0) if total_full else 0.0

    return {
        "empty": total_empty,
        "wip": total_wip,
        "full": total_full,
        "kanban": total_kanban,
        "scrap": total_scrap,
        "total_containers": total_containers,
        "tonnage": total_tonnage,
        "avg_kg_per_full": avg_kg_per_full,
        "kanban_pct": kanban_pct,
        "site_tonnage": site_tonnage,
        "site_full": site_full,
    }


def _kpi_card_excel(
    ws,
    row: int,
    col: int,
    width: int,
    label: str,
    value: str,
    sub: str = "",
    tone: str = "blue",
) -> None:
    """Render a KPI card into a 3-row × ``width``-col block.

    Row layout:
      row N      → label band (small, navy bg, white text)
      row N+1..2 → value (large bold) merged with sub line below
    """
    accent = {
        "blue":   "1F3A8A",
        "green":  "047857",
        "amber":  "B45309",
        "rose":   "BE123C",
        "slate":  "334155",
    }.get(tone, "1F3A8A")

    # Label band
    end_col = col + width - 1
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = Font(bold=True, color="FFFFFF", size=10)
    label_cell.fill = PatternFill("solid", fgColor=accent)
    label_cell.alignment = Alignment(
        horizontal="left", vertical="center", indent=1,
    )
    ws.row_dimensions[row].height = 20

    # Value row
    ws.merge_cells(
        start_row=row + 1, start_column=col,
        end_row=row + 1, end_column=end_col,
    )
    value_cell = ws.cell(row=row + 1, column=col, value=value)
    value_cell.font = Font(bold=True, size=20, color="0F172A")
    value_cell.alignment = Alignment(
        horizontal="left", vertical="center", indent=1,
    )
    value_cell.fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[row + 1].height = 34

    # Sub row
    ws.merge_cells(
        start_row=row + 2, start_column=col,
        end_row=row + 2, end_column=end_col,
    )
    sub_cell = ws.cell(row=row + 2, column=col, value=sub)
    sub_cell.font = Font(italic=True, size=9, color="64748B")
    sub_cell.alignment = Alignment(
        horizontal="left", vertical="center", indent=1,
    )
    sub_cell.fill = PatternFill("solid", fgColor="FFFFFF")
    ws.row_dimensions[row + 2].height = 16

    # Border sadece merged 3 satırın top-left hücrelerinde — daha
    # önce iç hücrelere ws.cell() ile dokunuyorduk, bu 'empty numeric
    # cells inside merged range' üretip Excel'i sorunlu buluyordu.
    thin = Side(style="thin", color="CBD5E1")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    label_cell.border = box
    value_cell.border = box
    sub_cell.border = box


def _link_button_excel(
    ws,
    row: int,
    col: int,
    width: int,
    height: int,
    label: str,
    target_sheet: str,
    target_cell: str = "A1",
    font_size: int = 11,
) -> None:
    """Render a navy 'button' cell with an internal hyperlink.

    openpyxl gerçek form-control buton üretemiyor, ama hücreyi
    butona benzer stillendirip ``#'Sheet'!cell`` formatlı iç
    bağlantı veriyoruz. ``target_cell`` ile aynı sheet içindeki
    belirli satıra da link verebiliriz.
    """
    end_col = col + width - 1
    end_row = row + height - 1
    if width > 1 or height > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=end_row, end_column=end_col,
        )
    cell = ws.cell(row=row, column=col, value=label)
    cell.font = Font(bold=True, size=font_size, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="94A3B8")
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True,
    )
    # openpyxl: string form '#Sheet!A1' TargetMode='External' üretiyor,
    # Excel bunu sorunlu iç link olarak görüp dosya açılırken uyarı
    # veriyor. Hyperlink objesi location= ile TargetMode boş kalıyor
    # (internal), Excel temiz açıyor.
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=f"'{target_sheet}'!{target_cell}",
        display=label,
    )

    # Sadece top-left hücresine border — merged range içindeki iç
    # hücrelere ws.cell() ile dokunmak Excel'in beğenmediği 'empty
    # numeric cells inside merged range' üretiyordu. Merged cell fill
    # zaten top-left'ten türetiliyor.
    thin = Side(style="thin", color="64748B")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fmt_int_tr(n) -> str:
    try:
        return f"{int(round(float(n))):,}".replace(",", ".")
    except (TypeError, ValueError):
        return str(n)


def _fmt_dec_tr(n, digits: int = 1) -> str:
    try:
        v = float(n)
    except (TypeError, ValueError):
        return str(n)
    formatted = f"{v:,.{digits}f}"
    return formatted.replace(",", "X").replace(".", ",").replace("X", ".")



def _build_renk_kirilim_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    """Detail per (department × color) for the selected week."""
    ws = wb.create_sheet("Renk Kırılımı")

    headers = [
        "Hafta", "Üretim Yeri", "Bölüm", "Renk",
        "Boş", "Proseste", "Dolu", "Dolu İçindeki Kanban", "Hurdaya Ayrılacak",
        "Toplam Konteyner", "Durum",
        "Giren Kullanıcı", "Sayım Tarihi", "Sayım Saati", "Gönderim Zamanı",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    totals = {"empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0, "bdh": 0}

    for idx, row in enumerate(rows, start=2):
        is_late = row.get("Durum") == "late_submitted"
        zebra = (idx % 2 == 0) and not is_late
        fill = _LATE_FILL if is_late else (_ZEBRA_FILL if zebra else None)

        bos_v = int(row.get("Boş") or 0)
        wip_v = int(row.get("Proseste") or 0)
        dolu_v = int(row.get("Dolu") or 0)
        kanban_v = int(row.get("Kanban") or 0)
        hurda_v = int(row.get("Hurda") or 0)
        bdh_v = bos_v + wip_v + dolu_v + hurda_v

        values = [
            row.get("Hafta", ""),
            row.get("Üretim Yeri", ""), row.get("Bölüm", ""), row.get("Renk", ""),
            row.get("Boş"), row.get("Proseste"),
            row.get("Dolu"), row.get("Kanban"), row.get("Hurda"),
            bdh_v,
            _STATUS_LABEL.get(row.get("Durum"), row.get("Durum", "")),
            row.get("Giren Kullanıcı", ""),
            row.get("Sayım Tarihi", ""), row.get("Sayım Saati", ""),
            _fmt_ts(row.get("Gönderim Zamanı")),
        ]
        ws.append(values)
        totals["empty"] += bos_v
        totals["wip"] += wip_v
        totals["full"] += dolu_v
        totals["kanban"] += kanban_v
        totals["scrap"] += hurda_v
        totals["bdh"] += bdh_v

        for col_idx in range(1, len(values) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.border = _BORDER
            if fill:
                cell.fill = fill
            # Sayısal sütunlar: 5=Boş, 6=WIP, 7=Dolu, 8=Kanban, 9=Hurda, 10=Toplam
            if col_idx in (5, 6, 7, 8, 9, 10):
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            elif col_idx == 11:  # Durum
                cell.alignment = _CENTER
                if is_late:
                    cell.font = Font(bold=True, color="92400E")
            elif col_idx == 1:  # Hafta
                cell.alignment = _CENTER
            else:
                cell.alignment = _LEFT

    if rows:
        total_row_idx = ws.max_row + 1
        ws.append([
            "TOPLAM", "", "", "",
            totals["empty"], totals["wip"], totals["full"],
            totals["kanban"], totals["scrap"], totals["bdh"],
            "", "", "", "", "",
        ])
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
            cell.border = _BORDER
            if col_idx in (5, 6, 7, 8, 9, 10):
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            elif col_idx == 1:
                cell.alignment = _RIGHT
            else:
                cell.alignment = _LEFT

    _autofit(ws, headers)


def _build_uretim_yeri_kirilim_sheet(
    wb: Workbook, rows: list[dict[str, Any]]
) -> dict[tuple[str, str], dict[str, Any]]:
    """Sheet 2: per-(site, department) aggregate.

    Returns the aggregate dict so the per-site sheet can reuse it
    without re-scanning the rows.
    """
    ws = wb.create_sheet("Üretim Yeri Kırılımı")

    headers = [
        "Üretim Yeri", "Bölüm",
        "Boş", "Proseste", "Dolu", "Dolu İçindeki Kanban", "Hurdaya ayrılacak",
        "Toplam Konteyner", "Toplam Tonaj",
        "Durum", "Giren Kullanıcı", "Sayım Gönderim Zamanı",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    aggregates: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        key = (row.get("Üretim Yeri", ""), row.get("Bölüm", ""))
        agg = aggregates.setdefault(key, {
            "empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0,
            "tonnage": row.get("Gerçekleşen Tonaj"),
            "status": row.get("Durum"),
            "user": row.get("Giren Kullanıcı", ""),
            "submitted_at": row.get("Gönderim Zamanı"),
        })
        agg["empty"] += int(row.get("Boş") or 0)
        agg["wip"] += int(row.get("Proseste") or 0)
        agg["full"] += int(row.get("Dolu") or 0)
        agg["kanban"] += int(row.get("Kanban") or 0)
        agg["scrap"] += int(row.get("Hurda") or 0)

    totals = {"empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0, "bdh": 0, "tonnage": 0.0}

    for idx, ((site, dept), agg) in enumerate(
        sorted(
            aggregates.items(),
            key=lambda kv: (_site_sort_key(kv[0][0]), kv[0][1]),
        ),
        start=2,
    ):
        is_late = agg["status"] == "late_submitted"
        fill = _LATE_FILL if is_late else (_ZEBRA_FILL if idx % 2 == 0 else None)
        bdh = (
            int(agg["empty"] or 0) + int(agg["wip"] or 0)
            + int(agg["full"] or 0) + int(agg["scrap"] or 0)
        )
        ton = agg["tonnage"]
        values = [
            site, dept,
            agg["empty"], agg["wip"], agg["full"], agg["kanban"], agg["scrap"],
            bdh, ton,
            _STATUS_LABEL.get(agg["status"], agg["status"] or ""),
            agg["user"],
            _fmt_ts(agg["submitted_at"]),
        ]
        ws.append(values)
        totals["empty"] += int(agg["empty"] or 0)
        totals["wip"] += int(agg["wip"] or 0)
        totals["full"] += int(agg["full"] or 0)
        totals["kanban"] += int(agg["kanban"] or 0)
        totals["scrap"] += int(agg["scrap"] or 0)
        totals["bdh"] += bdh
        if ton is not None:
            try:
                totals["tonnage"] += float(ton)
            except (TypeError, ValueError):
                pass

        for col_idx in range(1, len(values) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.border = _BORDER
            if fill:
                cell.fill = fill
            # Sayısal: 3=Boş, 4=WIP, 5=Dolu, 6=Kanban, 7=Hurda, 8=Toplam, 9=Tonaj
            if col_idx in (3, 4, 5, 6, 7, 8, 9):
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            elif col_idx == 10:  # Durum
                cell.alignment = _CENTER
                if is_late:
                    cell.font = Font(bold=True, color="92400E")
            else:
                cell.alignment = _LEFT

    if aggregates:
        total_row_idx = ws.max_row + 1
        ws.append([
            "TOPLAM", "",
            totals["empty"], totals["wip"], totals["full"],
            totals["kanban"], totals["scrap"],
            totals["bdh"], totals["tonnage"],
            "", "", "",
        ])
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
            cell.border = _BORDER
            if col_idx in (3, 4, 5, 6, 7, 8, 9):
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            elif col_idx == 1:
                cell.alignment = _RIGHT
            else:
                cell.alignment = _LEFT

    _autofit(ws, headers)
    return aggregates


def _build_uretim_yeri_ozeti_sheet(
    wb: Workbook, dept_aggs: dict[tuple[str, str], dict[str, Any]]
) -> None:
    """Sheet 3: per-site aggregate with percentage and ton/dolu KPI."""
    ws = wb.create_sheet("Üretim Yeri Özeti")
    headers = [
        "Üretim Yeri",
        "Boş", "Proseste", "Dolu", "Dolu içindeki Kanban", "Hurdaya ayrılacak",
        "Toplam Konteyner", "Toplam (%)",
        "Toplam Tonaj", "Dolu Konteyner Başına Yük (ton/konteyner)",
    ]
    ws.append(headers)
    # Uzun başlıklar (Dolu içindeki Kanban, Hurdaya ayrılacak, Toplam
    # Konteyner, Dolu Konteyner Başına Yük …) tek satıra sığmıyor —
    # wrap_text + yükseltilmiş satır yüksekliği ile iki satıra kırılıyor.
    _style_header_row(ws, len(headers), wrap_text=True, row_height=42)

    site_aggs: dict[str, dict[str, Any]] = {}
    for (site, _dept), agg in dept_aggs.items():
        s = site_aggs.setdefault(site, {
            "empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0, "tonnage": 0.0,
        })
        s["empty"] += int(agg["empty"] or 0)
        s["wip"] += int(agg.get("wip") or 0)
        s["full"] += int(agg["full"] or 0)
        s["kanban"] += int(agg["kanban"] or 0)
        s["scrap"] += int(agg["scrap"] or 0)
        if agg["tonnage"] is not None:
            try:
                s["tonnage"] += float(agg["tonnage"])
            except (TypeError, ValueError):
                pass

    grand_total_bdh = sum(
        s["empty"] + s["wip"] + s["full"] + s["scrap"] for s in site_aggs.values()
    )

    totals = {"empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0, "bdh": 0, "tonnage": 0.0}
    for idx, (site, s) in enumerate(
        sorted(site_aggs.items(), key=lambda kv: _site_sort_key(kv[0])),
        start=2,
    ):
        zebra = _ZEBRA_FILL if idx % 2 == 0 else None
        bdh = s["empty"] + s["wip"] + s["full"] + s["scrap"]
        pct = (bdh / grand_total_bdh) if grand_total_bdh else 0  # stored as fraction
        ton_per = (s["tonnage"] / s["full"]) if s["full"] else 0

        values = [
            site,
            s["empty"], s["wip"], s["full"], s["kanban"], s["scrap"],
            bdh, pct,
            s["tonnage"], ton_per,
        ]
        ws.append(values)
        totals["empty"] += s["empty"]
        totals["wip"] += s["wip"]
        totals["full"] += s["full"]
        totals["kanban"] += s["kanban"]
        totals["scrap"] += s["scrap"]
        totals["bdh"] += bdh
        totals["tonnage"] += s["tonnage"]

        for col_idx in range(1, len(values) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.border = _BORDER
            if zebra:
                cell.fill = zebra
            # Sayısal: 2=Boş, 3=WIP, 4=Dolu, 5=Kanban, 6=Hurda, 7=Toplam
            if col_idx in (2, 3, 4, 5, 6, 7):
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            elif col_idx == 8:  # Toplam %
                cell.alignment = _RIGHT
                cell.number_format = "0.0%"
            elif col_idx == 9:  # Toplam Tonaj
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            elif col_idx == 10:  # ton/konteyner
                cell.alignment = _RIGHT
                cell.number_format = "0.00"
            else:
                cell.alignment = _LEFT

    if site_aggs:
        total_row_idx = ws.max_row + 1
        ton_per_total = (totals["tonnage"] / totals["full"]) if totals["full"] else 0
        ws.append([
            "TOPLAM",
            totals["empty"], totals["wip"], totals["full"],
            totals["kanban"], totals["scrap"],
            totals["bdh"], 1.0,
            totals["tonnage"], ton_per_total,
        ])
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
            cell.border = _BORDER
            if col_idx in (2, 3, 4, 5, 6, 7, 9):
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            elif col_idx == 8:
                cell.alignment = _RIGHT
                cell.number_format = "0.0%"
            elif col_idx == 10:
                cell.alignment = _RIGHT
                cell.number_format = "0.00"
            elif col_idx == 1:
                cell.alignment = _RIGHT
            else:
                cell.alignment = _LEFT

    _autofit(ws, headers)
    # Wrap_text yaptığımız uzun başlıkların sütun genişliğini
    # _autofit header uzunluğuna göre 40+'a fişlemesin diye
    # küçültüyoruz — başlık iki satıra kırılıyor, hücre değerleri
    # zaten kısa.
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["J"].width = 14


def _build_renk_ozeti_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    """Sheet 4: per-color aggregate for the selected week."""
    ws = wb.create_sheet("Renk Özeti")
    headers = [
        "Renk", "Boş", "Proseste", "Dolu", "Dolu içindeki Kanban",
        "Hurdaya ayrılacak", "Toplam Konteyner",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers), wrap_text=True, row_height=42)

    color_aggs: dict[str, dict[str, int]] = {}
    color_order: list[str] = []
    for row in rows:
        color = row.get("Renk", "") or ""
        if color not in color_aggs:
            color_aggs[color] = {"empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0}
            color_order.append(color)
        c = color_aggs[color]
        c["empty"] += int(row.get("Boş") or 0)
        c["wip"] += int(row.get("Proseste") or 0)
        c["full"] += int(row.get("Dolu") or 0)
        c["kanban"] += int(row.get("Kanban") or 0)
        c["scrap"] += int(row.get("Hurda") or 0)

    totals = {"empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0, "bdh": 0}
    for idx, color in enumerate(color_order, start=2):
        c = color_aggs[color]
        bdh = c["empty"] + c["wip"] + c["full"] + c["scrap"]
        values = [color, c["empty"], c["wip"], c["full"], c["kanban"], c["scrap"], bdh]
        ws.append(values)
        totals["empty"] += c["empty"]
        totals["wip"] += c["wip"]
        totals["full"] += c["full"]
        totals["kanban"] += c["kanban"]
        totals["scrap"] += c["scrap"]
        totals["bdh"] += bdh

        zebra = _ZEBRA_FILL if idx % 2 == 0 else None
        for col_idx in range(1, len(values) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.border = _BORDER
            if zebra:
                cell.fill = zebra
            if col_idx >= 2:
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            else:
                cell.alignment = _LEFT

    if color_order:
        total_row_idx = ws.max_row + 1
        ws.append([
            "TOPLAM",
            totals["empty"], totals["wip"], totals["full"], totals["kanban"],
            totals["scrap"], totals["bdh"],
        ])
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
            cell.border = _BORDER
            if col_idx >= 2:
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            else:
                cell.alignment = _RIGHT

    _autofit(ws, headers)
    # Wrap_text uyguladığımız uzun başlıklar için sütun genişliği
    # kapatılıyor — autofit aksi takdirde başlık uzunluğunu baz alıp
    # 20+ veriyor.
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 13


# ---------------------------------------------------------------------------
# Multi-week aggregation (used by ÖZET + Karşılaştırma sheets)
# ---------------------------------------------------------------------------

def _aggregate_all_weeks(
    all_rows: list[dict[str, Any]],
    manual_aggs: list[dict[str, Any]] | None = None,
) -> tuple[
    dict[str, dict[str, float]],           # weekly_totals
    dict[str, dict[str, dict[str, float]]], # weekly_site
    dict[str, dict[str, int]],             # weekly_color (color → toplam konteyner)
    list[str],                              # ordered color list (discovery order)
    set[str],                              # manual_only_weeks
]:
    """Aggregate the long-format rows into structures useful for charts.

    Tonnage is captured once per (week, site, department) because each color
    row of the same submission carries the same submission-level tonnage —
    summing it per color would multiply by color count.

    ``manual_aggs`` lets the caller fold historical / pre-system data
    (rows from the ``manual_site_aggregates`` table) into the weekly +
    site aggregates without polluting the per-color or per-department
    structures (which have no manual-data counterpart).
    """
    weekly_totals: dict[str, dict[str, float]] = {}
    weekly_site: dict[str, dict[str, dict[str, float]]] = {}
    weekly_color: dict[str, dict[str, int]] = {}
    color_order: list[str] = []
    seen_colors: set[str] = set()
    # Weeks that have any count_submission row (i.e. real per-color data).
    # Anything not in here but folded in via manual_aggs is "manual only"
    # and should be excluded from charts/sheets that depend on the
    # per-color or per-tonnage dimension.
    submission_weeks: set[str] = set()

    tonnage_seen: set[tuple[str, str, str]] = set()

    for r in all_rows:
        week = r.get("Hafta") or ""
        site = r.get("Üretim Yeri") or ""
        dept = r.get("Bölüm") or ""
        color = r.get("Renk") or ""

        if week:
            submission_weeks.add(week)

        empty_v = int(r.get("Boş") or 0)
        wip_v = int(r.get("Proseste") or 0)
        full_v = int(r.get("Dolu") or 0)
        kanban_v = int(r.get("Kanban") or 0)
        scrap_v = int(r.get("Hurda") or 0)
        # Toplam Konteyner = Boş + WIP + Dolu + Hurda (yeni tanım).
        bdh_v = empty_v + wip_v + full_v + scrap_v

        wt = weekly_totals.setdefault(week, {
            "empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0,
            "bdh": 0, "tonnage": 0.0,
        })
        wt["empty"] += empty_v
        wt["wip"] += wip_v
        wt["full"] += full_v
        wt["kanban"] += kanban_v
        wt["scrap"] += scrap_v
        wt["bdh"] += bdh_v

        ws_agg = weekly_site.setdefault(week, {}).setdefault(site, {
            "empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0,
            "bdh": 0, "tonnage": 0.0,
        })
        ws_agg["empty"] += empty_v
        ws_agg["wip"] += wip_v
        ws_agg["full"] += full_v
        ws_agg["kanban"] += kanban_v
        ws_agg["scrap"] += scrap_v
        ws_agg["bdh"] += bdh_v

        if color and color not in seen_colors:
            seen_colors.add(color)
            color_order.append(color)
        wc = weekly_color.setdefault(week, {})
        wc[color] = wc.get(color, 0) + bdh_v

        key = (week, site, dept)
        if key not in tonnage_seen:
            tonnage_seen.add(key)
            t = r.get("Gerçekleşen Tonaj")
            if t is not None:
                try:
                    t_f = float(t)
                    wt["tonnage"] += t_f
                    ws_agg["tonnage"] += t_f
                except (TypeError, ValueError):
                    pass

    # Fold in the manual aggregates (no color / dept granularity); these
    # only touch ``weekly_totals`` and ``weekly_site``. Per-color charts
    # skip these weeks naturally because there are no color rows.
    if manual_aggs:
        for m in manual_aggs:
            week = m.get("week_iso") or ""
            site = m.get("site") or ""
            if not week or not site:
                continue
            empty_v = int(m.get("empty") or 0)
            full_v = int(m.get("full") or 0)
            scrap_v = int(m.get("scrap") or 0)
            wip_v = int(m.get("wip") or 0)
            bdh_v = empty_v + wip_v + full_v + scrap_v
            tonnage_v = m.get("tonnage")

            wt = weekly_totals.setdefault(week, {
                "empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0,
                "bdh": 0, "tonnage": 0.0,
            })
            wt["empty"] += empty_v
            wt["wip"] += wip_v
            wt["full"] += full_v
            wt["scrap"] += scrap_v
            wt["bdh"] += bdh_v

            ws_agg = weekly_site.setdefault(week, {}).setdefault(site, {
                "empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0,
                "bdh": 0, "tonnage": 0.0,
            })
            ws_agg["empty"] += empty_v
            ws_agg["wip"] += wip_v
            ws_agg["full"] += full_v
            ws_agg["scrap"] += scrap_v
            ws_agg["bdh"] += bdh_v

            if tonnage_v is not None:
                try:
                    t_f = float(tonnage_v)
                    wt["tonnage"] += t_f
                    ws_agg["tonnage"] += t_f
                except (TypeError, ValueError):
                    pass

    manual_only_weeks = {
        week for week in weekly_totals.keys() if week not in submission_weeks
    }

    return (
        weekly_totals, weekly_site, weekly_color, color_order,
        manual_only_weeks,
    )


# ---------------------------------------------------------------------------
# ÖZET — charts across all weeks
# ---------------------------------------------------------------------------

# Display order used by the color-breakdown chart. Anything outside this
# list is appended at the end in discovery order.
_DEFINED_COLOR_ORDER = [
    "Mavi", "Turuncu", "Yeşil", "Gri", "MS Vida", "Sarı",
    "Mor", "Kırmızı", "Siyah",
]


# Production-site display order for charts where üretim yerleri sit on the
# X axis (charts 3 and 4). Anything not listed falls in alphabetically at
# the end so a newly-added site is still visible without a code change.
_SITE_ORDER = [
    "Norm Cıvata İzmir",
    "Norm Cıvata Salihli",
    "Norm Somun İzmir",
    "Norm Somun Salihli",
    "Uysal İzmir",
    "Uysal Salihli",
    "MS Vida",
    "Nedu",
    "Sac Şekillendirme",
    "Sıcak Dövme",
    "Norm Holding",
]


def _site_sort_key(site: str) -> tuple[int, str]:
    """Sort by ``_SITE_ORDER`` position; unknown sites fall to the end."""
    try:
        return (_SITE_ORDER.index(site), site)
    except ValueError:
        return (len(_SITE_ORDER), site)


def _short_week(week_iso: str) -> str:
    """Strip the year prefix from an ISO week code for chart axis use.

    ``2026-W18`` → ``W18``. Falls back to the original string if the
    input doesn't match the expected shape.
    """
    if not week_iso:
        return ""
    if "-W" in week_iso:
        try:
            _, suffix = week_iso.split("-W", 1)
            return f"W{suffix}"
        except ValueError:
            return week_iso
    return week_iso


def _clean_axis(axis) -> None:
    """Common axis styling: visible labels, no major gridlines."""
    axis.delete = False
    axis.majorGridlines = None
    axis.majorTickMark = "out"


def _horizontal_axis_title(text: str) -> Title:
    """Build an axis title rendered horizontally and anchored near the
    top of the axis instead of the default vertically-centered rotated
    layout.

    Excel's default Y-axis title is rotated 90° and centered along the
    axis. This helper forces ``rot=0`` (horizontal) and positions the
    title near the top edge so it reads naturally above the topmost
    Y-axis tick.
    """
    body_pr = RichTextProperties(rot=0, vert="horz")
    char_props = CharacterProperties(b=True, sz=1000)
    para_props = ParagraphProperties(defRPr=char_props)
    run = RegularTextRun(t=text)
    para = Paragraph(pPr=para_props, r=[run])
    rt = RichText(bodyPr=body_pr, p=[para])
    tx = Text(rich=rt)
    layout = Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.05, xMode="edge", yMode="edge",
        )
    )
    return Title(tx=tx, layout=layout, overlay=False)


def _make_chart_title(text: str, size_pt: int = 13) -> Title:
    """Standard chart title: bold, ``size_pt`` (default 13pt), top-center.

    Replaces the implicit string-to-Title conversion openpyxl does when
    you assign ``chart.title = "..."``; with that path the title is
    rendered at the default 18pt non-bold. We pass an explicit Title
    instead so every chart in the workbook gets a consistent 13pt bold
    heading.
    """
    body_pr = RichTextProperties(rot=0, vert="horz")
    char_props = CharacterProperties(b=True, sz=size_pt * 100)
    para_props = ParagraphProperties(defRPr=char_props)
    run = RegularTextRun(t=text)
    para = Paragraph(pPr=para_props, r=[run])
    rt = RichText(bodyPr=body_pr, p=[para])
    tx = Text(rich=rt)
    return Title(tx=tx, overlay=False)


def _apply_chart_frame(chart, *, color: str = "475569", width_emu: int = 22000):
    """Give the chart a thick outer frame so it pops on the sheet.

    ``width_emu``: 1 pt ≈ 12700 EMU. 22000 EMU ≈ 1.75pt.
    """
    try:
        gp = GraphicalProperties()
        gp.line = LineProperties(w=width_emu, solidFill=color)
        chart.graphical_properties = gp
    except Exception:
        # openpyxl version differences — just skip the frame styling.
        pass


def _white_bold_label_props() -> RichText:
    """Text properties for stacked-segment labels — bold white so the
    number stands out against the colored fill of the bar."""
    char_props = CharacterProperties(
        b=True, sz=1000, solidFill="FFFFFF",
    )
    para_props = ParagraphProperties(defRPr=char_props)
    return RichText(
        bodyPr=RichTextProperties(rot=0, vert="horz"),
        p=[Paragraph(pPr=para_props)],
    )


def _bold_large_label_props(
    size_pt: int = 11, color: str | None = None,
) -> RichText:
    """Text properties for the Toplam overlay label — bold and slightly
    bigger than the body so the total reads as the headline number for
    each cluster / stack. ``color`` (hex without #) forces explicit text
    color; varsayılan Excel auto-renk."""
    kwargs: dict[str, Any] = {"b": True, "sz": size_pt * 100}
    if color:
        kwargs["solidFill"] = color
    char_props = CharacterProperties(**kwargs)
    para_props = ParagraphProperties(defRPr=char_props)
    return RichText(
        bodyPr=RichTextProperties(rot=0, vert="horz"),
        p=[Paragraph(pPr=para_props)],
    )


def _hide_overlay_from_legend(chart, overlay) -> None:
    """Drop the overlay chart's series from the host chart's legend.

    After ``chart += overlay``, the overlay's series sit AFTER the
    host chart's series in the legend ordering — but
    ``chart.series`` keeps returning only the host's count. Adding
    the overlay's series count finds the real legend index. We
    assign a *list* because openpyxl's Legend.legendEntry is a
    Sequence; assigning a single LegendEntry silently no-ops.
    """
    primary = len(chart.series)
    secondary = len(getattr(overlay, "series", []))
    total = primary + secondary
    if total == 0:
        return
    try:
        chart.legend.legendEntry = [
            LegendEntry(idx=total - 1, delete=True),
        ]
    except Exception:
        pass


def _top_left_chart_title(text: str) -> Title:
    """Chart title anchored at the top-left corner instead of top-center.

    Useful when data labels above the bars would otherwise overlap the
    centered title (e.g. chart 4 has tall stacks with Toplam labels on
    top — centered title collides with the tallest site's label).
    """
    body_pr = RichTextProperties(rot=0, vert="horz")
    char_props = CharacterProperties(b=True, sz=1100)
    para_props = ParagraphProperties(defRPr=char_props)
    run = RegularTextRun(t=text)
    para = Paragraph(pPr=para_props, r=[run])
    rt = RichText(bodyPr=body_pr, p=[para])
    tx = Text(rich=rt)
    layout = Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02, xMode="edge", yMode="edge",
        )
    )
    return Title(tx=tx, layout=layout, overlay=False)


def _end_x_axis_title(text: str) -> Title:
    """Build an X-axis title anchored near the right edge of the chart
    (i.e. at the *end* of the X axis) instead of below its center.

    Matches ``_horizontal_axis_title`` styling so X and Y titles look
    visually consistent across charts.
    """
    body_pr = RichTextProperties(rot=0, vert="horz")
    char_props = CharacterProperties(b=True, sz=1000)
    para_props = ParagraphProperties(defRPr=char_props)
    run = RegularTextRun(t=text)
    para = Paragraph(pPr=para_props, r=[run])
    rt = RichText(bodyPr=body_pr, p=[para])
    tx = Text(rich=rt)
    layout = Layout(
        manualLayout=ManualLayout(
            x=0.93, y=0.93, xMode="edge", yMode="edge",
        )
    )
    return Title(tx=tx, layout=layout, overlay=False)


def _value_only_labels(
    position: str,
    num_format: str = "[$-tr-TR]#,##0",
    txPr: RichText | None = None,
) -> DataLabelList:
    """Data labels that show ONLY the numeric value.

    openpyxl/Excel sometimes display series name and category name
    alongside the value when those flags are left unset (interpreted
    as 'show by default'). Setting every other show* flag to False
    keeps the label to just the number.

    The ``num_format`` is written to the ``<c:dLbls><c:numFmt/>``
    element, but Excel still falls back to the source-cell format
    when ``sourceLinked`` defaults to true (and openpyxl can't write
    ``sourceLinked="0"`` here — its DataLabelList only accepts the
    format code as a string). To make the format actually take
    effect, ALSO call ``cell.number_format = num_format`` on the
    backing ``_veri`` sheet cells. The string passed here is
    redundant-but-harmless fallback.

    ``txPr`` lets the caller override the label text style (color,
    bold, size) — used for the white bold labels inside stacked
    bars and the slightly larger bold Toplam label.
    """
    return DataLabelList(
        showVal=True,
        showCatName=False,
        showSerName=False,
        showLegendKey=False,
        showPercent=False,
        showBubbleSize=False,
        dLblPos=position,
        numFmt=num_format,
        txPr=txPr,
    )


def _build_ozet_charts_sheet(
    wb: Workbook,
    all_rows: list[dict[str, Any]],
    manual_aggs: list[dict[str, Any]] | None = None,
) -> None:
    """Sheet 5 (ÖZET): four charts only.

    Backing data tables live in a ``veryHidden`` ``_veri`` sheet so the
    visible ÖZET sheet shows the title and the four charts and nothing
    else. Hidden-but-referenced cells continue to feed the charts.
    """
    ws = wb.create_sheet("Grafikler")

    # Title banner — same styling as Dashboard sheet's banner so the two
    # tabs read as siblings.
    ws.merge_cells("A1:X1")
    title_cell = ws["A1"]
    title_cell.value = "Konteyner Dashboard — Haftalık Analiz"
    title_cell.font = Font(bold=True, size=20, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F3A8A")
    title_cell.alignment = Alignment(
        horizontal="left", vertical="center", indent=1,
    )
    ws.row_dimensions[1].height = 38

    if not all_rows and not manual_aggs:
        ws["A4"] = "Henüz veri yok — sayım girildikçe burada grafikler oluşacak."
        return

    # Hidden helper sheet: 'veryHidden' keeps it out of the unhide menu so
    # casual users can't stumble onto the raw data tables.
    data_ws = wb.create_sheet("_veri")
    data_ws.sheet_state = "veryHidden"

    (
        weekly_totals, weekly_site, weekly_color, color_order,
        manual_only_weeks,
    ) = _aggregate_all_weeks(all_rows, manual_aggs)
    weeks = sorted(weekly_totals.keys())
    # Weeks that have full per-color / tonnage data — used by charts 2,
    # 3, and 4 plus the comparison sheet so manual-only entries (Boş +
    # Dolu only, no tonnage, no color) don't render as misleading
    # zeros.
    full_weeks = [w for w in weeks if w not in manual_only_weeks]
    # Custom site order — see _SITE_ORDER. Anything not on that list
    # falls through alphabetically at the end.
    all_sites = sorted(
        {s for sd in weekly_site.values() for s in sd.keys()},
        key=_site_sort_key,
    )
    last_3_weeks = full_weeks[-3:] if len(full_weeks) >= 3 else full_weeks[:]
    latest_week = full_weeks[-1] if full_weeks else None

    # Tesis Detayı bölümünün anchor satırları — chart 3 ve chart 5'in
    # yanındaki üretim yeri butonları buraya link veriyor. Sadece
    # tesis blokları render ediliyor (bölüm drilldown kaldırıldı).
    _n_weeks = len(full_weeks)
    _main_block_rows = max(_n_weeks + 3, 17) + 4
    detail_section_header_row = 305
    detail_blocks_start_row = detail_section_header_row + 2

    site_anchors: dict[str, int] = {
        site: detail_blocks_start_row + i * _main_block_rows
        for i, site in enumerate(all_sites)
    }

    # Latest-week snapshot used by per-chart KPI side panels.
    if latest_week:
        latest_rows_only = [
            r for r in all_rows if r.get("Hafta") == latest_week
        ]
        latest_kpis = _compute_week_kpis(latest_rows_only)

        # Subtitle: which week the KPIs reflect.
        ws.merge_cells("A2:X2")
        sub_cell = ws["A2"]
        sub_cell.value = (
            f"Son Hafta: {_short_week(latest_week)} — "
            f"Her grafiğin yanındaki kartlar o grafiğin KPI'larını gösterir."
        )
        sub_cell.font = Font(italic=True, size=11, color="475569")
        sub_cell.alignment = Alignment(
            horizontal="left", vertical="center", indent=1,
        )
        ws.row_dimensions[2].height = 20
    else:
        latest_kpis = None

    # Sol blok (sütun 1..20) grafiklerin ferahça oturduğu alan, sağ
    # blok (21..24) yandaki KPI panelleri / üretim yeri butonları.
    # 24'ten sonrası kullanılmıyor.
    for c in range(1, 25):
        ws.column_dimensions[get_column_letter(c)].width = 11
    # Tesis Detayı tablolarındaki 5 sütun: Hafta / Yarı Mamul Tonajı /
    # Boş Konteyner / Dolu Konteyner / Dolu Konteyner Tonajı. Uzun
    # başlıklar wrap_text + genişletilmiş kolonla iki satıra kırılıyor.
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15

    # Section header before the trend charts band (row 4).
    if latest_week:
        ws.merge_cells("A4:X4")
        sec_cell = ws["A4"]
        sec_cell.value = "Haftalık Trendler"
        sec_cell.font = Font(bold=True, size=13, color="1F3A8A")
        sec_cell.fill = PatternFill("solid", fgColor="E2E8F0")
        sec_cell.alignment = Alignment(
            horizontal="left", vertical="center", indent=1,
        )
        ws.row_dimensions[4].height = 24

    # ================================================================
    # Chart 1 — Clustered column: weekly total split by category
    #   X = weeks
    #   Series order: Boş, Dolu, Dolu İçindeki Kanban, Hurda (matches
    #   the per-week tables elsewhere in the workbook)
    # ================================================================
    # Yeni 'Toplam Konteyner' tanımı: Boş + WIP + Dolu + Hurda. Kanban
    # hâlâ Dolu'nun alt kümesi olduğu için bu grafikte ayrı bir
    # kategori değil — 4 ayrık kategori stackleniyor.
    t1_col = 1
    t1_headers = ["Hafta", "Boş", "Proseste", "Dolu", "Hurdaya Ayrılacak", "Toplam"]
    for j, h in enumerate(t1_headers):
        data_ws.cell(row=1, column=t1_col + j, value=h)
    for i, w in enumerate(weeks):
        wt = weekly_totals[w]
        # Short label on the X axis ('W18' instead of '2026-W18') keeps
        # the category labels compact and readable across all charts.
        data_ws.cell(row=2 + i, column=t1_col, value=_short_week(w))
        # Apply Turkish thousand-separated format to data cells so the
        # chart data labels (which Excel pulls via sourceLinked=true)
        # inherit the format. See _value_only_labels() for rationale.
        for col_offset, key in enumerate(
            ["empty", "wip", "full", "scrap"], start=1
        ):
            cell = data_ws.cell(
                row=2 + i, column=t1_col + col_offset, value=wt.get(key, 0),
            )
            cell.number_format = "[$-tr-TR]#,##0"
        # Toplam = Boş + WIP + Dolu + Hurda. Aynı stack yüksekliği,
        # dolayısıyla overlay etiketi stack tepesine birebir oturur.
        total = (
            wt.get("empty", 0) + wt.get("wip", 0)
            + wt.get("full", 0) + wt.get("scrap", 0)
        )
        total_cell = data_ws.cell(
            row=2 + i, column=t1_col + 5, value=total
        )
        total_cell.number_format = "[$-tr-TR]#,##0"
    t1_last = 1 + len(weeks)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 2
    chart1.grouping = "stacked"
    chart1.overlap = 100
    chart1.title = _make_chart_title("Haftalık Toplam Konteyner Dağılımı")
    chart1.y_axis.title = _horizontal_axis_title("Konteyner Adedi")
    chart1.x_axis.title = _end_x_axis_title("Hafta")
    data_ref = Reference(
        data_ws,
        min_col=t1_col + 1, min_row=1,
        max_col=t1_col + 4, max_row=t1_last,
    )
    chart1.add_data(data_ref, titles_from_data=True)
    cats_ref = Reference(data_ws, min_col=t1_col, min_row=2, max_row=t1_last)
    chart1.set_categories(cats_ref)
    _clean_axis(chart1.x_axis)
    _clean_axis(chart1.y_axis)
    # Turkish thousand-separated format on Y-axis tick labels (5000 → 5.000)
    chart1.y_axis.numFmt = "[$-tr-TR]#,##0"
    # Per-segment data labels centered inside each stack segment — white
    # bold so they pop against the colored fill. Small segments may have
    # the label spill outside, where readability falls back to the chart
    # background (acceptable trade-off; user accepted this).
    chart1.dataLabels = _value_only_labels(
        "ctr", txPr=_white_bold_label_props(),
    )
    chart1.legend.position = "b"
    chart1.legend.overlay = False
    # Grafik 38 cm geniş, 10 cm yüksek — daha kompakt; yandaki KPI
    # paneli col U (21) civarında başlıyor, çakışmıyor.
    chart1.height = 10
    chart1.width = 38
    _apply_chart_frame(chart1)

    # Invisible "Toplam" line at the top of each stack — same value as
    # the stack height (B + WIP + D + H), so the data label lands
    # right at the stack top. Bigger bold label so the headline total
    # reads as the cluster summary.
    total_line = LineChart()
    total_ref = Reference(
        data_ws,
        min_col=t1_col + 5, min_row=1,
        max_col=t1_col + 5, max_row=t1_last,
    )
    total_line.add_data(total_ref, titles_from_data=True)
    total_line.set_categories(cats_ref)
    for s in total_line.series:
        gp = GraphicalProperties()
        gp.line = LineProperties(noFill=True)
        s.graphicalProperties = gp
        s.marker = Marker(symbol="none")
    total_line.dataLabels = _value_only_labels(
        "t", txPr=_bold_large_label_props(size_pt=12),
    )
    chart1 += total_line
    # Drop the 'Toplam' series from the legend (it's just an overlay
    # for the label; the bars below already cover the colored swatches).
    _hide_overlay_from_legend(chart1, total_line)

    # Hurda values are tiny compared to Boş / WIP / Dolu so the
    # centered white label overflows the colored segment onto the
    # white chart background and becomes unreadable. Drop Hurda's
    # per-segment labels entirely; the Toplam overlay above the
    # stack plus the data tables in the other sheets still surface
    # the Hurda number when the user needs it.
    # Bar series order matches the stack bottom-up: 0=Boş, 1=WIP,
    # 2=Dolu, 3=Hurdaya Ayrılacak.
    if len(chart1.series) >= 4:
        chart1.series[3].dLbls = DataLabelList(delete=True)

    chart1_anchor_row = 5
    ws.add_chart(chart1, f"A{chart1_anchor_row}")

    # ================================================================
    # Chart 1B — Yarı Mamul Stok Tonajı ve Boş Konteyner İlişkisi
    #   Bar (tonaj — sol Y ekseni turuncu) + Line overlay (boş — sağ
    #   Y ekseni yeşil, secondary axis). Her hafta yarı mamul stoğu
    #   ile boş konteyner sayısı nasıl birlikte değişiyor.
    #
    #   ÖNEMLİ: overlay.y_axis.crossAx = 10 (chart_rel'in catAx axId'i)
    #   set edilmezse Excel XML'de crossAx eksik oluyor ve dosya
    #   'Çizim şekli hatalı' uyarısı veriyor. crossAx=10 ile temiz.
    # ================================================================
    t_rel_col = 60  # sağ tarafta boş kolon — çakışma yok
    data_ws.cell(row=1, column=t_rel_col, value="Hafta")
    data_ws.cell(
        row=1, column=t_rel_col + 1, value="Yarı Mamul Tonajı",
    )
    data_ws.cell(row=1, column=t_rel_col + 2, value="Boş Konteyner")
    for i, w in enumerate(full_weeks):
        wt = weekly_totals[w]
        data_ws.cell(row=2 + i, column=t_rel_col, value=_short_week(w))
        c_ton = data_ws.cell(
            row=2 + i, column=t_rel_col + 1,
            value=float(wt.get("tonnage", 0.0)),
        )
        c_ton.number_format = "[$-tr-TR]#,##0"
        c_emp = data_ws.cell(
            row=2 + i, column=t_rel_col + 2,
            value=int(wt.get("empty", 0)),
        )
        c_emp.number_format = "[$-tr-TR]#,##0"
    t_rel_last = 1 + len(full_weeks)

    chart_rel = BarChart()
    chart_rel.type = "col"
    chart_rel.style = 2
    chart_rel.title = _make_chart_title(
        "Yarı Mamul Stok Tonajı ve Boş Konteyner İlişkisi"
    )
    chart_rel.y_axis.title = _horizontal_axis_title("Tonaj (t)")
    chart_rel.x_axis.title = _end_x_axis_title("Hafta")
    if full_weeks:
        chart_rel.add_data(
            Reference(
                data_ws,
                min_col=t_rel_col + 1, min_row=1,
                max_col=t_rel_col + 1, max_row=t_rel_last,
            ),
            titles_from_data=True,
        )
        chart_rel.set_categories(
            Reference(
                data_ws, min_col=t_rel_col, min_row=2, max_row=t_rel_last,
            )
        )
    _clean_axis(chart_rel.x_axis)
    _clean_axis(chart_rel.y_axis)
    chart_rel.y_axis.numFmt = "[$-tr-TR]#,##0"
    chart_rel.y_axis.scaling.min = 0
    chart_rel.gapWidth = 30
    for s in chart_rel.series:
        gp = GraphicalProperties(solidFill="EA580C")
        s.graphicalProperties = gp
    chart_rel.dataLabels = _value_only_labels(
        "t", "[$-tr-TR]#,##0",
        txPr=_bold_large_label_props(size_pt=10, color="0F172A"),
    )

    empty_line_overlay = LineChart()
    if full_weeks:
        empty_line_overlay.add_data(
            Reference(
                data_ws,
                min_col=t_rel_col + 2, min_row=1,
                max_col=t_rel_col + 2, max_row=t_rel_last,
            ),
            titles_from_data=True,
        )
        empty_line_overlay.set_categories(
            Reference(
                data_ws, min_col=t_rel_col, min_row=2, max_row=t_rel_last,
            )
        )
    for s in empty_line_overlay.series:
        s.marker = Marker(symbol="circle", size=8)
        gp = GraphicalProperties()
        gp.line = LineProperties(solidFill="047857", w=28000)
        s.graphicalProperties = gp
    empty_line_overlay.dataLabels = _value_only_labels(
        "t", "[$-tr-TR]#,##0",
        txPr=_bold_large_label_props(size_pt=10, color="047857"),
    )
    # Secondary Y ekseni — axId farklı (200), crossAx = catAx (10),
    # crosses='max' (sağ tarafta). crossAx eksik olursa Excel red
    # ediyor.
    empty_line_overlay.y_axis.axId = 200
    empty_line_overlay.y_axis.crossAx = 10
    empty_line_overlay.y_axis.crosses = "max"
    empty_line_overlay.y_axis.title = _horizontal_axis_title(
        "Boş Konteyner Adedi"
    )
    empty_line_overlay.y_axis.numFmt = "[$-tr-TR]#,##0"
    empty_line_overlay.y_axis.scaling.min = 0
    chart_rel += empty_line_overlay

    chart_rel.legend.position = "b"
    chart_rel.legend.overlay = False
    chart_rel.height = 10
    chart_rel.width = 38
    _apply_chart_frame(chart_rel)
    chart_rel_anchor_row = 28
    ws.add_chart(chart_rel, f"A{chart_rel_anchor_row}")

    # Side KPI panel for chart 1 — categories shown in the stack
    # (Toplam = stack top, Boş / Proseste / Dolu / Hurda are the
    # stacked bars). All values reflect the latest week.
    if latest_kpis:
        _kpi_card_excel(
            ws, row=chart1_anchor_row, col=21, width=4,
            label="Toplam Konteyner",
            value=_fmt_int_tr(latest_kpis["total_containers"]),
            sub=f"Son hafta: {_short_week(latest_week)}",
            tone="slate",
        )
        _kpi_card_excel(
            ws, row=chart1_anchor_row + 4, col=21, width=4,
            label="Boş",
            value=_fmt_int_tr(latest_kpis["empty"]),
            sub="Kullanılabilir kasa",
            tone="green",
        )
        _kpi_card_excel(
            ws, row=chart1_anchor_row + 8, col=21, width=4,
            label="Proseste",
            value=_fmt_int_tr(latest_kpis["wip"]),
            sub="İşlem görüyor",
            tone="amber",
        )
        _kpi_card_excel(
            ws, row=chart1_anchor_row + 12, col=21, width=4,
            label="Dolu (Kanban dahil)",
            value=_fmt_int_tr(latest_kpis["full"]),
            sub=(
                f"Kanban: {_fmt_int_tr(latest_kpis['kanban'])} "
                f"(%{_fmt_dec_tr(latest_kpis['kanban_pct'])})"
            ),
            tone="blue",
        )
        # Hurda kartı — Toplam = Boş + Proseste + Dolu + Hurda olduğu
        # için Hurda da panelde görünmeli; aksi takdirde dört kartın
        # toplamı Toplam'a denk gelmiyor.
        _kpi_card_excel(
            ws, row=chart1_anchor_row + 16, col=21, width=4,
            label="Hurdaya Ayrılacak",
            value=_fmt_int_tr(latest_kpis["scrap"]),
            sub="Kullanım dışı",
            tone="rose",
        )

    # ================================================================
    # Chart 2 — Line chart: overall ton/Dolu per week
    #   Single series across all weeks with markers + data labels.
    #   Data column layout: each row is one week so the chart resolves
    #   to ONE series with N points (not N series with 1 point each).
    # ================================================================
    t2_col = 7  # well clear of table 1
    data_ws.cell(row=1, column=t2_col, value="Hafta")
    data_ws.cell(row=1, column=t2_col + 1, value="Ton / Dolu Konteyner")
    # Use full_weeks (excludes manual-only entries) so W17 etc. don't
    # appear as misleading zeros on the ton/Dolu chart.
    for i, w in enumerate(full_weeks):
        wt = weekly_totals[w]
        ton_per = (wt["tonnage"] / wt["full"]) if wt["full"] else None
        data_ws.cell(row=2 + i, column=t2_col, value=_short_week(w))
        cell = data_ws.cell(row=2 + i, column=t2_col + 1, value=ton_per)
        cell.number_format = "[$-tr-TR]#,##0.00"
    t2_last = 1 + len(full_weeks)

    chart2 = LineChart()
    chart2.style = 2
    chart2.title = _make_chart_title("Dolu Konteyner Başına Tonaj")
    chart2.y_axis.title = _horizontal_axis_title("Ton / Dolu Konteyner")
    chart2.x_axis.title = _end_x_axis_title("Hafta")
    data_ref = Reference(
        data_ws,
        min_col=t2_col + 1, min_row=1,
        max_col=t2_col + 1, max_row=t2_last,
    )
    chart2.add_data(data_ref, titles_from_data=True)
    cats_ref = Reference(data_ws, min_col=t2_col, min_row=2, max_row=t2_last)
    chart2.set_categories(cats_ref)
    _clean_axis(chart2.x_axis)
    _clean_axis(chart2.y_axis)
    # ton/Dolu is fractional → two decimal places on both the Y-axis
    # tick labels and the per-point data labels.
    chart2.y_axis.numFmt = "[$-tr-TR]#,##0.00"
    # Tight Y-axis band (0.20 → 0.90) — values typically cluster in this
    # range, and forcing the floor/ceiling surfaces week-on-week variance
    # that Excel's auto-scale would otherwise flatten.
    chart2.y_axis.scaling.min = 0.20
    chart2.y_axis.scaling.max = 0.90
    # Veri etiketleri kalın + 13 pt + koyu lacivert — auto-renk
    # ile soluk gri tonda kalıyordu, sabit koyu renkle (0F172A)
    # hem kalın hem yüksek kontrastla okunuyor.
    chart2.dataLabels = _value_only_labels(
        "t", "[$-tr-TR]#,##0.00",
        txPr=_bold_large_label_props(size_pt=13, color="0F172A"),
    )
    for series in chart2.series:
        series.marker = Marker(symbol="circle", size=7)
        # Force a single solid line color (navy blue) so the line reads
        # as one continuous series instead of a default per-segment
        # auto-gradient.
        gp = GraphicalProperties()
        gp.line = LineProperties(solidFill="1F3A8A", w=22000)
        series.graphicalProperties = gp
    chart2.legend = None  # single series — legend is just noise
    chart2.height = 10
    chart2.width = 38
    _apply_chart_frame(chart2)
    # Chart 2 sits below chart 1 panel.
    chart2_anchor_row = 120
    ws.add_chart(chart2, f"A{chart2_anchor_row}")

    # ================================================================
    # Chart 3 — Clustered column: ton/Dolu per site for last 3 weeks
    #   X = production sites
    #   Series = last 3 weeks (one cluster of 3 bars per site)
    # ================================================================
    t3_col = 10
    data_ws.cell(row=1, column=t3_col, value="Üretim Yeri")
    for j, w in enumerate(last_3_weeks):
        # Series headers become legend entries → short week label.
        data_ws.cell(row=1, column=t3_col + 1 + j, value=_short_week(w))
    for i, site in enumerate(all_sites):
        data_ws.cell(row=2 + i, column=t3_col, value=site)
        for j, w in enumerate(last_3_weeks):
            sd = weekly_site.get(w, {}).get(site)
            val = (sd["tonnage"] / sd["full"]) if (sd and sd["full"]) else None
            cell = data_ws.cell(row=2 + i, column=t3_col + 1 + j, value=val)
            cell.number_format = "[$-tr-TR]#,##0.00"
    t3_last = 1 + len(all_sites)

    chart3 = BarChart()
    chart3.type = "col"
    chart3.style = 2
    chart3.grouping = "clustered"
    chart3.title = _make_chart_title(
        "Dolu Konteyner Başına Tonaj — Üretim Yeri Kırılımı (Son 3 Hafta)"
    )
    chart3.y_axis.title = _horizontal_axis_title("Ton / Dolu Konteyner")
    chart3.x_axis.title = _end_x_axis_title("Üretim Yeri")
    if last_3_weeks and all_sites:
        data_ref = Reference(
            data_ws,
            min_col=t3_col + 1, min_row=1,
            max_col=t3_col + len(last_3_weeks), max_row=t3_last,
        )
        chart3.add_data(data_ref, titles_from_data=True)
        cats_ref = Reference(data_ws, min_col=t3_col, min_row=2, max_row=t3_last)
        chart3.set_categories(cats_ref)
    _clean_axis(chart3.x_axis)
    _clean_axis(chart3.y_axis)
    chart3.y_axis.scaling.min = 0
    # ton/Dolu is fractional → two decimal places everywhere (Y-axis +
    # data labels).
    chart3.y_axis.numFmt = "[$-tr-TR]#,##0.00"
    chart3.dataLabels = _value_only_labels("outEnd", "[$-tr-TR]#,##0.00")
    chart3.legend.position = "b"
    chart3.legend.overlay = False
    # Wider than other charts: 11 sites × 3 weekly bars per site = 33
    # columns, so the chart needs the extra width to keep labels
    # legible without overlapping.
    chart3.height = 14
    chart3.width = 38
    _apply_chart_frame(chart3)
    # Section divider before the tesis-comparison charts.
    ws.merge_cells("A143:X143")
    sec3 = ws["A143"]
    sec3.value = "Tesis Karşılaştırma"
    sec3.font = Font(bold=True, size=13, color="1F3A8A")
    sec3.fill = PatternFill("solid", fgColor="E2E8F0")
    sec3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[143].height = 24
    chart3_anchor_row = 144
    ws.add_chart(chart3, f"A{chart3_anchor_row}")

    # KPI yerine: her üretim yeri için ayrı buton. Tıklanınca o
    # tesisin sayfanın alt kısmındaki Tesis Detayı bloğuna atlar
    # (ton/Dolu + Boş trend mini grafikleri).
    if all_sites:
        n_sites = len(all_sites)
        btn_h = 2
        chart_span_rows = 28  # chart 3 height 14 cm ≈ 28 satır
        total_btn_h = n_sites * btn_h
        start_offset = max(0, (chart_span_rows - total_btn_h) // 2)
        for i, site in enumerate(all_sites):
            _link_button_excel(
                ws,
                row=chart3_anchor_row + start_offset + i * btn_h,
                col=21, width=4, height=btn_h,
                label=site,
                target_sheet="Grafikler",
                target_cell=f"A{site_anchors[site]}",
                font_size=10,
            )

    # ================================================================
    # Chart 3B — Yarı Mamul Tonajı — Üretim Yeri Kırılımı (Son 3 Hafta)
    #   Chart 3 ile aynı yapı ama y = haftalık ham tonaj (t).
    # ================================================================
    t_tsite_col = 45  # chart_full_trend (40-41), wtonnage (42-43) uzağı
    data_ws.cell(row=1, column=t_tsite_col, value="Üretim Yeri")
    for j, w in enumerate(last_3_weeks):
        data_ws.cell(
            row=1, column=t_tsite_col + 1 + j, value=_short_week(w),
        )
    for i, site in enumerate(all_sites):
        data_ws.cell(row=2 + i, column=t_tsite_col, value=site)
        for j, w in enumerate(last_3_weeks):
            sd = weekly_site.get(w, {}).get(site)
            val = float(sd.get("tonnage", 0.0)) if sd else 0.0
            cell = data_ws.cell(
                row=2 + i, column=t_tsite_col + 1 + j, value=val,
            )
            cell.number_format = "[$-tr-TR]#,##0"
    t_tsite_last = 1 + len(all_sites)

    chart_tsite = BarChart()
    chart_tsite.type = "col"
    chart_tsite.style = 2
    chart_tsite.grouping = "clustered"
    chart_tsite.title = _make_chart_title(
        "Yarı Mamul Tonajı — Üretim Yeri Kırılımı (Son 3 Hafta)"
    )
    chart_tsite.y_axis.title = _horizontal_axis_title("Tonaj (t)")
    chart_tsite.x_axis.title = _end_x_axis_title("Üretim Yeri")
    if last_3_weeks and all_sites:
        chart_tsite.add_data(
            Reference(
                data_ws,
                min_col=t_tsite_col + 1, min_row=1,
                max_col=t_tsite_col + len(last_3_weeks),
                max_row=t_tsite_last,
            ),
            titles_from_data=True,
        )
        chart_tsite.set_categories(
            Reference(
                data_ws, min_col=t_tsite_col, min_row=2,
                max_row=t_tsite_last,
            )
        )
    _clean_axis(chart_tsite.x_axis)
    _clean_axis(chart_tsite.y_axis)
    chart_tsite.y_axis.numFmt = "[$-tr-TR]#,##0"
    chart_tsite.y_axis.scaling.min = 0
    chart_tsite.dataLabels = _value_only_labels("outEnd", "[$-tr-TR]#,##0")
    chart_tsite.legend.position = "b"
    chart_tsite.legend.overlay = False
    chart_tsite.height = 14
    chart_tsite.width = 38
    _apply_chart_frame(chart_tsite)
    chart_tsite_anchor_row = 175
    ws.add_chart(chart_tsite, f"A{chart_tsite_anchor_row}")

    # ================================================================
    # Chart 3.5 — Toplam Boş Konteyner — Haftalık Trend
    #   Tüm tesislerin haftalık toplam boş konteyner trendi. Tek
    #   seri, full_weeks boyunca.
    # ================================================================
    t_empty_col = t3_col + len(last_3_weeks) + 2  # chart 3 verisinden gap
    data_ws.cell(row=1, column=t_empty_col, value="Hafta")
    data_ws.cell(
        row=1, column=t_empty_col + 1, value="Toplam Boş Konteyner",
    )
    for i, w in enumerate(full_weeks):
        wt = weekly_totals[w]
        data_ws.cell(row=2 + i, column=t_empty_col, value=_short_week(w))
        cell = data_ws.cell(
            row=2 + i, column=t_empty_col + 1,
            value=int(wt.get("empty", 0)),
        )
        cell.number_format = "[$-tr-TR]#,##0"
    t_empty_last = 1 + len(full_weeks)

    chart_empty_trend = LineChart()
    chart_empty_trend.style = 2
    chart_empty_trend.title = _make_chart_title(
        "Toplam Boş Konteyner — Haftalık Trend"
    )
    chart_empty_trend.y_axis.title = _horizontal_axis_title(
        "Toplam Boş Konteyner"
    )
    chart_empty_trend.x_axis.title = _end_x_axis_title("Hafta")
    if full_weeks:
        data_ref = Reference(
            data_ws,
            min_col=t_empty_col + 1, min_row=1,
            max_col=t_empty_col + 1, max_row=t_empty_last,
        )
        chart_empty_trend.add_data(data_ref, titles_from_data=True)
        cats_ref_emp = Reference(
            data_ws, min_col=t_empty_col, min_row=2, max_row=t_empty_last,
        )
        chart_empty_trend.set_categories(cats_ref_emp)
    _clean_axis(chart_empty_trend.x_axis)
    _clean_axis(chart_empty_trend.y_axis)
    chart_empty_trend.y_axis.numFmt = "[$-tr-TR]#,##0"
    # Y ekseni sıfırdan başlasın — auto scale bazen negatif padding
    # ekleyip veri 0 iken çizgiyi eksiye düşüyormuş gibi gösteriyor.
    chart_empty_trend.y_axis.scaling.min = 0
    chart_empty_trend.dataLabels = _value_only_labels(
        "t", "[$-tr-TR]#,##0",
        txPr=_bold_large_label_props(size_pt=12, color="0F172A"),
    )
    for series in chart_empty_trend.series:
        series.marker = Marker(symbol="circle", size=7)
        gp = GraphicalProperties()
        # Boş ↔ yeşil — chart 5 (Tesis Bazlı Boş) ile tematik uyum.
        gp.line = LineProperties(solidFill="047857", w=22000)
        series.graphicalProperties = gp
    chart_empty_trend.legend = None
    chart_empty_trend.height = 10
    chart_empty_trend.width = 38
    _apply_chart_frame(chart_empty_trend)
    chart_empty_trend_anchor_row = 74
    ws.add_chart(chart_empty_trend, f"A{chart_empty_trend_anchor_row}")

    # ================================================================
    # Chart 3.6 — Toplam Dolu Konteyner — Haftalık Trend
    #   Boş grafiğinin dolu ikizi: aynı yapı (haftalık toplam), farklı
    #   seri. Data kolonu 40 — chart 4 (t4_col=18) ve chart 5
    #   (t5_col=~27) alanlarından uzakta, çakışmasın.
    # ================================================================
    t_full_col = 40
    data_ws.cell(row=1, column=t_full_col, value="Hafta")
    data_ws.cell(
        row=1, column=t_full_col + 1, value="Toplam Dolu Konteyner",
    )
    for i, w in enumerate(full_weeks):
        wt = weekly_totals[w]
        data_ws.cell(row=2 + i, column=t_full_col, value=_short_week(w))
        cell = data_ws.cell(
            row=2 + i, column=t_full_col + 1,
            value=int(wt.get("full", 0)),
        )
        cell.number_format = "[$-tr-TR]#,##0"
    t_full_last = 1 + len(full_weeks)

    chart_full_trend = LineChart()
    chart_full_trend.style = 2
    chart_full_trend.title = _make_chart_title(
        "Toplam Dolu Konteyner — Haftalık Trend"
    )
    chart_full_trend.y_axis.title = _horizontal_axis_title(
        "Toplam Dolu Konteyner"
    )
    chart_full_trend.x_axis.title = _end_x_axis_title("Hafta")
    if full_weeks:
        data_ref = Reference(
            data_ws,
            min_col=t_full_col + 1, min_row=1,
            max_col=t_full_col + 1, max_row=t_full_last,
        )
        chart_full_trend.add_data(data_ref, titles_from_data=True)
        cats_ref_full = Reference(
            data_ws, min_col=t_full_col, min_row=2, max_row=t_full_last,
        )
        chart_full_trend.set_categories(cats_ref_full)
    _clean_axis(chart_full_trend.x_axis)
    _clean_axis(chart_full_trend.y_axis)
    chart_full_trend.y_axis.numFmt = "[$-tr-TR]#,##0"
    chart_full_trend.y_axis.scaling.min = 0
    chart_full_trend.dataLabels = _value_only_labels(
        "t", "[$-tr-TR]#,##0",
        txPr=_bold_large_label_props(size_pt=12, color="0F172A"),
    )
    for series in chart_full_trend.series:
        series.marker = Marker(symbol="circle", size=7)
        gp = GraphicalProperties()
        # Dolu ↔ lacivert — chart 1 stack'inde Dolu segmenti ile uyumlu.
        gp.line = LineProperties(solidFill="1F3A8A", w=22000)
        series.graphicalProperties = gp
    chart_full_trend.legend = None
    chart_full_trend.height = 10
    chart_full_trend.width = 38
    _apply_chart_frame(chart_full_trend)
    chart_full_trend_anchor_row = 97
    ws.add_chart(chart_full_trend, f"A{chart_full_trend_anchor_row}")

    # ================================================================
    # Chart 1.5 — Haftalık Toplam Yarı Mamul Tonajı (LineChart)
    #   Tüm tesislerin haftalık ham tonaj toplamının zamansal trendi.
    # ================================================================
    t_wtonnage_col = 42  # chart_full_trend (40-41) alanının uzağı
    data_ws.cell(row=1, column=t_wtonnage_col, value="Hafta")
    data_ws.cell(
        row=1, column=t_wtonnage_col + 1, value="Toplam Tonaj (t)",
    )
    for i, w in enumerate(full_weeks):
        wt = weekly_totals[w]
        data_ws.cell(row=2 + i, column=t_wtonnage_col, value=_short_week(w))
        cell = data_ws.cell(
            row=2 + i, column=t_wtonnage_col + 1,
            value=float(wt.get("tonnage", 0.0)),
        )
        cell.number_format = "[$-tr-TR]#,##0"
    t_wtonnage_last = 1 + len(full_weeks)

    chart_wtonnage_trend = LineChart()
    chart_wtonnage_trend.style = 2
    chart_wtonnage_trend.title = _make_chart_title(
        "Haftalık Toplam Yarı Mamul Tonajı"
    )
    chart_wtonnage_trend.y_axis.title = _horizontal_axis_title("Tonaj (t)")
    chart_wtonnage_trend.x_axis.title = _end_x_axis_title("Hafta")
    if full_weeks:
        chart_wtonnage_trend.add_data(
            Reference(
                data_ws,
                min_col=t_wtonnage_col + 1, min_row=1,
                max_col=t_wtonnage_col + 1, max_row=t_wtonnage_last,
            ),
            titles_from_data=True,
        )
        chart_wtonnage_trend.set_categories(
            Reference(
                data_ws, min_col=t_wtonnage_col, min_row=2,
                max_row=t_wtonnage_last,
            )
        )
    _clean_axis(chart_wtonnage_trend.x_axis)
    _clean_axis(chart_wtonnage_trend.y_axis)
    chart_wtonnage_trend.y_axis.numFmt = "[$-tr-TR]#,##0"
    chart_wtonnage_trend.y_axis.scaling.min = 0
    chart_wtonnage_trend.dataLabels = _value_only_labels(
        "t", "[$-tr-TR]#,##0",
        txPr=_bold_large_label_props(size_pt=12, color="0F172A"),
    )
    for series in chart_wtonnage_trend.series:
        series.marker = Marker(symbol="circle", size=7)
        gp = GraphicalProperties()
        # Turuncu — tonaj temalı çizgi
        gp.line = LineProperties(solidFill="EA580C", w=22000)
        series.graphicalProperties = gp
    chart_wtonnage_trend.legend = None
    chart_wtonnage_trend.height = 10
    chart_wtonnage_trend.width = 38
    _apply_chart_frame(chart_wtonnage_trend)
    chart_wtonnage_trend_anchor_row = 51
    ws.add_chart(chart_wtonnage_trend, f"A{chart_wtonnage_trend_anchor_row}")

    # ================================================================
    # Chart 4 — Stacked column: color breakdown per site (latest week)
    #   X = production sites
    #   Stacks = colors in defined order
    #   Each color series painted with its brand hex
    # ================================================================
    # Aggregate per (site, color) for the latest week.
    site_color_latest: dict[str, dict[str, int]] = {}
    if latest_week:
        for r in all_rows:
            if r.get("Hafta") != latest_week:
                continue
            site = r.get("Üretim Yeri") or ""
            color = r.get("Renk") or ""
            cnt = (
                int(r.get("Boş") or 0)
                + int(r.get("Dolu") or 0)
                + int(r.get("Hurda") or 0)
            )
            site_color_latest.setdefault(site, {})
            site_color_latest[site][color] = (
                site_color_latest[site].get(color, 0) + cnt
            )

    # Color order: defined order first, then any extras in discovery order.
    chart4_colors = (
        [c for c in _DEFINED_COLOR_ORDER if c in color_order]
        + [c for c in color_order if c not in _DEFINED_COLOR_ORDER]
    )
    # Custom production-site order on the X axis.
    chart4_sites = sorted(site_color_latest.keys(), key=_site_sort_key)

    t4_col = 18
    data_ws.cell(row=1, column=t4_col, value="Üretim Yeri")
    for j, color in enumerate(chart4_colors):
        data_ws.cell(row=1, column=t4_col + 1 + j, value=color)
    # Column after the last color holds the per-site stack total used by
    # the invisible 'Toplam' overlay line.
    t4_total_col = t4_col + 1 + len(chart4_colors)
    data_ws.cell(row=1, column=t4_total_col, value="Toplam")
    for i, site in enumerate(chart4_sites):
        data_ws.cell(row=2 + i, column=t4_col, value=site)
        sc = site_color_latest.get(site, {})
        for j, color in enumerate(chart4_colors):
            cell = data_ws.cell(
                row=2 + i, column=t4_col + 1 + j, value=sc.get(color, 0)
            )
            cell.number_format = "[$-tr-TR]#,##0"
        total = sum(sc.values())
        total_cell = data_ws.cell(row=2 + i, column=t4_total_col, value=total)
        total_cell.number_format = "[$-tr-TR]#,##0"
    t4_last = 1 + len(chart4_sites)

    chart4 = BarChart()
    chart4.type = "col"
    chart4.style = 2
    chart4.grouping = "stacked"
    chart4.overlap = 100
    title_suffix = f" ({_short_week(latest_week)})" if latest_week else ""
    # Standard top-center bold title — chart 4 is now offset to the
    # right of column A (anchored at G), so the title-vs-label
    # collision the top-left layout was originally solving no longer
    # applies (the chart has its own left margin now).
    chart4.title = _make_chart_title(
        f"Üretim Yerleri Renk Dağılımı{title_suffix}"
    )
    chart4.y_axis.title = _horizontal_axis_title("Konteyner Adedi")
    chart4.x_axis.title = _end_x_axis_title("Üretim Yeri")
    if chart4_colors and chart4_sites:
        data_ref = Reference(
            data_ws,
            min_col=t4_col + 1, min_row=1,
            max_col=t4_col + len(chart4_colors), max_row=t4_last,
        )
        chart4.add_data(data_ref, titles_from_data=True)
        cats_ref4 = Reference(
            data_ws, min_col=t4_col, min_row=2, max_row=t4_last,
        )
        chart4.set_categories(cats_ref4)
        for series, color_name in zip(chart4.series, chart4_colors):
            hex_code = _COLOR_HEX.get(color_name, "94A3B8")
            try:
                _set_bar_series_color(series, hex_code)
            except Exception:
                pass
    _clean_axis(chart4.x_axis)
    _clean_axis(chart4.y_axis)
    chart4.y_axis.numFmt = "[$-tr-TR]#,##0"
    # No per-segment data labels: with 6 stacked colors × 11 sites the
    # values pile on top of each other. We do show a single 'Toplam'
    # label above each stack via an invisible overlay line (below).
    chart4.legend.position = "b"
    chart4.legend.overlay = False
    chart4.height = 16
    chart4.width = 38
    _apply_chart_frame(chart4)

    # Invisible 'Toplam' line that sits at the top of each stack and
    # carries the total-count data label (bigger / bold so it reads as
    # the per-site headline). Mirrors the chart-1 pattern.
    if chart4_colors and chart4_sites:
        chart4_total_line = LineChart()
        total_ref4 = Reference(
            data_ws,
            min_col=t4_total_col, min_row=1,
            max_col=t4_total_col, max_row=t4_last,
        )
        chart4_total_line.add_data(total_ref4, titles_from_data=True)
        chart4_total_line.set_categories(cats_ref4)
        for s in chart4_total_line.series:
            gp = GraphicalProperties()
            gp.line = LineProperties(noFill=True)
            s.graphicalProperties = gp
            s.marker = Marker(symbol="none")
        chart4_total_line.dataLabels = _value_only_labels(
            "t", txPr=_bold_large_label_props(size_pt=12),
        )
        chart4 += chart4_total_line
        # Drop 'Toplam' from the legend swatches.
        _hide_overlay_from_legend(chart4, chart4_total_line)

    # Section divider before the color breakdown chart.
    ws.merge_cells("A268:X268")
    sec4 = ws["A268"]
    sec4.value = "Renk Dağılımı"
    sec4.font = Font(bold=True, size=13, color="1F3A8A")
    sec4.fill = PatternFill("solid", fgColor="E2E8F0")
    sec4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[268].height = 24
    chart4_anchor_row = 269
    ws.add_chart(chart4, f"A{chart4_anchor_row}")

    # Side KPI panel for chart 4 — color totals across all sites in
    # the latest week.
    if latest_week:
        color_totals_lw: dict[str, int] = {}
        for r in all_rows:
            if r.get("Hafta") != latest_week:
                continue
            color = r.get("Renk") or ""
            cnt = (
                int(r.get("Boş") or 0)
                + int(r.get("WIP") or 0)
                + int(r.get("Dolu") or 0)
                + int(r.get("Hurda") or 0)
            )
            color_totals_lw[color] = color_totals_lw.get(color, 0) + cnt

        if color_totals_lw:
            sorted_colors = sorted(
                color_totals_lw.items(), key=lambda kv: kv[1], reverse=True,
            )
            most = sorted_colors[0]
            least = sorted_colors[-1]
            grand_total = sum(color_totals_lw.values())

            _kpi_card_excel(
                ws, row=chart4_anchor_row, col=21, width=4,
                label="Toplam Konteyner",
                value=_fmt_int_tr(grand_total),
                sub=f"{_short_week(latest_week)} — tüm renkler",
                tone="slate",
            )
            _kpi_card_excel(
                ws, row=chart4_anchor_row + 4, col=21, width=4,
                label="En Çok Konteyner",
                value=_fmt_int_tr(most[1]),
                sub=f"{most[0]}",
                tone="green",
            )
            _kpi_card_excel(
                ws, row=chart4_anchor_row + 8, col=21, width=4,
                label="En Az Konteyner",
                value=_fmt_int_tr(least[1]),
                sub=f"{least[0]}",
                tone="rose",
            )
            _kpi_card_excel(
                ws, row=chart4_anchor_row + 12, col=21, width=4,
                label="Farklı Renk Sayısı",
                value=_fmt_int_tr(len(color_totals_lw)),
                sub="Latest week aktif renkler",
                tone="amber",
            )

    # ================================================================
    # Chart 5 — Tesis Bazlı Boş Konteyner (Son 3 Hafta)
    #   Mirrors chart 3's layout: X axis = production sites, three
    #   clustered bars per site (last 3 weeks), value labels on top.
    #   The 'Boş' counterpart to chart 3's ton/Dolu view — per
    #   Selim Bey's brief, this is the early-warning view that tells
    #   us 'Salih Cıvata'nın boş'u 5000 → 4000 → 2800' before
    #   production stops.
    # ================================================================
    t5_col = t4_total_col + 2  # leave a gap past chart 4's data
    data_ws.cell(row=1, column=t5_col, value="Üretim Yeri")
    for j, w in enumerate(last_3_weeks):
        data_ws.cell(row=1, column=t5_col + 1 + j, value=_short_week(w))
    for i, site in enumerate(all_sites):
        data_ws.cell(row=2 + i, column=t5_col, value=site)
        for j, w in enumerate(last_3_weeks):
            sd = weekly_site.get(w, {}).get(site)
            empty_val = int(sd["empty"]) if sd else 0
            cell = data_ws.cell(
                row=2 + i, column=t5_col + 1 + j, value=empty_val,
            )
            cell.number_format = "[$-tr-TR]#,##0"
    t5_last = 1 + len(all_sites)

    chart5 = BarChart()
    chart5.type = "col"
    chart5.style = 2
    chart5.grouping = "clustered"
    chart5.title = _make_chart_title(
        "Tesis Bazlı Boş Konteyner (Son 3 Hafta)"
    )
    chart5.y_axis.title = _horizontal_axis_title("Boş Konteyner Adedi")
    chart5.x_axis.title = _end_x_axis_title("Üretim Yeri")
    if last_3_weeks and all_sites:
        data_ref = Reference(
            data_ws,
            min_col=t5_col + 1, min_row=1,
            max_col=t5_col + len(last_3_weeks), max_row=t5_last,
        )
        chart5.add_data(data_ref, titles_from_data=True)
        cats_ref5 = Reference(
            data_ws, min_col=t5_col, min_row=2, max_row=t5_last,
        )
        chart5.set_categories(cats_ref5)
    _clean_axis(chart5.x_axis)
    _clean_axis(chart5.y_axis)
    chart5.y_axis.numFmt = "[$-tr-TR]#,##0"
    chart5.y_axis.scaling.min = 0
    # Value labels on top of each bar — same outEnd styling as chart 3
    # so the two clustered charts read consistently. Bold for emphasis.
    chart5.dataLabels = _value_only_labels(
        "outEnd", txPr=_bold_large_label_props(size_pt=10),
    )
    chart5.legend.position = "b"
    chart5.legend.overlay = False
    chart5.height = 14
    chart5.width = 38
    _apply_chart_frame(chart5)
    chart5_anchor_row = 206
    ws.add_chart(chart5, f"A{chart5_anchor_row}")

    # KPI yerine: chart 3 ile aynı şekilde her üretim yeri için
    # buton. Aynı detay bloklarına link veriyor — kullanıcı her iki
    # paneldan da aynı tesise erişebiliyor.
    if all_sites:
        n_sites = len(all_sites)
        btn_h = 2
        chart_span_rows = 28
        total_btn_h = n_sites * btn_h
        start_offset = max(0, (chart_span_rows - total_btn_h) // 2)
        for i, site in enumerate(all_sites):
            _link_button_excel(
                ws,
                row=chart5_anchor_row + start_offset + i * btn_h,
                col=21, width=4, height=btn_h,
                label=site,
                target_sheet="Grafikler",
                target_cell=f"A{site_anchors[site]}",
                font_size=10,
            )

    # ================================================================
    # Chart 5B — Tesis Bazlı Dolu Konteyner (Son 3 Hafta)
    #   Chart 5'in dolu ikizi — aynı yapı, farklı metrik.
    # ================================================================
    t5f_col = 51  # chart_tsite (45-48) alanının uzağı
    data_ws.cell(row=1, column=t5f_col, value="Üretim Yeri")
    for j, w in enumerate(last_3_weeks):
        data_ws.cell(
            row=1, column=t5f_col + 1 + j, value=_short_week(w),
        )
    for i, site in enumerate(all_sites):
        data_ws.cell(row=2 + i, column=t5f_col, value=site)
        for j, w in enumerate(last_3_weeks):
            sd = weekly_site.get(w, {}).get(site)
            full_v = int(sd.get("full", 0)) if sd else 0
            cell = data_ws.cell(
                row=2 + i, column=t5f_col + 1 + j, value=full_v,
            )
            cell.number_format = "[$-tr-TR]#,##0"
    t5f_last = 1 + len(all_sites)

    chart5f = BarChart()
    chart5f.type = "col"
    chart5f.style = 2
    chart5f.grouping = "clustered"
    chart5f.title = _make_chart_title(
        "Tesis Bazlı Dolu Konteyner (Son 3 Hafta)"
    )
    chart5f.y_axis.title = _horizontal_axis_title("Dolu Konteyner Adedi")
    chart5f.x_axis.title = _end_x_axis_title("Üretim Yeri")
    if last_3_weeks and all_sites:
        chart5f.add_data(
            Reference(
                data_ws,
                min_col=t5f_col + 1, min_row=1,
                max_col=t5f_col + len(last_3_weeks), max_row=t5f_last,
            ),
            titles_from_data=True,
        )
        chart5f.set_categories(
            Reference(
                data_ws, min_col=t5f_col, min_row=2, max_row=t5f_last,
            )
        )
    _clean_axis(chart5f.x_axis)
    _clean_axis(chart5f.y_axis)
    chart5f.y_axis.numFmt = "[$-tr-TR]#,##0"
    chart5f.y_axis.scaling.min = 0
    chart5f.dataLabels = _value_only_labels(
        "outEnd", txPr=_bold_large_label_props(size_pt=10),
    )
    chart5f.legend.position = "b"
    chart5f.legend.overlay = False
    chart5f.height = 14
    chart5f.width = 38
    _apply_chart_frame(chart5f)
    chart5f_anchor_row = 237
    ws.add_chart(chart5f, f"A{chart5f_anchor_row}")

    # ================================================================
    # Tesis Detayı — per-site weekly trend blocks (chart 3 ve chart 5
    # yan butonları buradaki bloklara link veriyor)
    # ================================================================
    if all_sites and full_weeks:
        ws.merge_cells(
            f"A{detail_section_header_row}:X{detail_section_header_row}"
        )
        sec_d = ws.cell(
            row=detail_section_header_row, column=1,
            value="Tesis Detayı — Üretim Yeri Bazlı Haftalık Trend",
        )
        sec_d.font = Font(bold=True, size=13, color="1F3A8A")
        sec_d.fill = PatternFill("solid", fgColor="E2E8F0")
        sec_d.alignment = Alignment(
            horizontal="left", vertical="center", indent=1,
        )
        ws.row_dimensions[detail_section_header_row].height = 24

        for site in all_sites:
            block_row = site_anchors[site]

            # Banner
            ws.merge_cells(
                start_row=block_row, start_column=1,
                end_row=block_row, end_column=24,
            )
            bnr = ws.cell(
                row=block_row, column=1, value=f"{site} — Haftalık Trend",
            )
            bnr.font = Font(bold=True, size=14, color="FFFFFF")
            bnr.fill = PatternFill("solid", fgColor="1F3A8A")
            bnr.alignment = Alignment(
                horizontal="left", vertical="center", indent=1,
            )
            ws.row_dimensions[block_row].height = 28

            # Üst kısma dön bağlantısı
            back = ws.cell(
                row=block_row + 1, column=1, value="◀ Üst kısma dön",
            )
            back.hyperlink = Hyperlink(
                ref=back.coordinate,
                location="'Grafikler'!A1",
                display="◀ Üst kısma dön",
            )
            back.font = Font(
                bold=True, color="1F3A8A", size=11, underline="single",
            )
            back.alignment = Alignment(horizontal="left", vertical="center")

            # Tablo başlığı — sıra: Hafta / Yarı Mamul Tonajı / Boş /
            # Dolu / Dolu Konteyner Tonajı. Uzun başlıklar wrap_text
            # ile iki satıra kırılıyor.
            hdr_row = block_row + 2
            wrap_center = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
            )
            for j, h in enumerate(
                ["Hafta", "Yarı Mamul Tonajı", "Boş Konteyner",
                 "Dolu Konteyner", "Dolu Konteyner Tonajı"],
                start=1,
            ):
                c = ws.cell(row=hdr_row, column=j, value=h)
                c.fill = _HEADER_FILL
                c.font = _HEADER_FONT
                c.alignment = wrap_center
                c.border = _BORDER
            ws.row_dimensions[hdr_row].height = 40

            # Veri satırları — Yarı Mamul Tonajı ve Dolu Konteyner
            # Tonajı ikisi de submission-level 'tonnage'; kullanıcı
            # her ikisini de ham tonaj olarak gösteriyor.
            for k, w in enumerate(full_weeks):
                r = hdr_row + 1 + k
                sd = weekly_site.get(w, {}).get(site)
                empty_v = int(sd.get("empty", 0)) if sd else None
                full_v = int(sd.get("full", 0)) if sd else None
                ton_v = float(sd.get("tonnage", 0.0)) if sd else None
                c1 = ws.cell(row=r, column=1, value=_short_week(w))
                c1.alignment = _LEFT
                c1.border = _BORDER
                c2 = ws.cell(row=r, column=2, value=ton_v)
                c2.alignment = _RIGHT
                c2.number_format = "#,##0"
                c2.border = _BORDER
                c3 = ws.cell(row=r, column=3, value=empty_v)
                c3.alignment = _RIGHT
                c3.number_format = "#,##0"
                c3.border = _BORDER
                c4 = ws.cell(row=r, column=4, value=full_v)
                c4.alignment = _RIGHT
                c4.number_format = "#,##0"
                c4.border = _BORDER
                c5 = ws.cell(row=r, column=5, value=ton_v)
                c5.alignment = _RIGHT
                c5.number_format = "#,##0"
                c5.border = _BORDER

            table_end_row = hdr_row + len(full_weeks)

            # Bu tesisin haftalık ham değerleri — mini chart Y ekseni
            # scaling'i için. Auto scale, değerler dar bir bantta
            # (örn. Uysal İzmir 100-105) yer aldığında integer format
            # ile tick label'larını aynı sayıya yuvarlıyor: 100, 100,
            # 100... Manuel min=0 + max=%20 padding'le rahat okunuyor.
            site_ton_vals = [
                float(
                    weekly_site.get(w, {}).get(site, {}).get("tonnage", 0.0)
                    or 0
                )
                for w in full_weeks
            ]
            site_empty_vals = [
                int(
                    weekly_site.get(w, {}).get(site, {}).get("empty", 0)
                    or 0
                )
                for w in full_weeks
            ]
            site_full_vals = [
                int(
                    weekly_site.get(w, {}).get(site, {}).get("full", 0)
                    or 0
                )
                for w in full_weeks
            ]

            # 4 mini chart yan yana — sütun sırasıyla aynı: Yarı Mamul
            # Tonajı | Boş | Dolu | Dolu Konteyner Tonajı. Cols F/J/N/R.
            def _mini(
                anchor: str, title: str, src_col: int,
                num_fmt: str, line_color: str, series_vals: list[float],
            ) -> None:
                ch = LineChart()
                ch.title = _make_chart_title(title)
                ch.add_data(
                    Reference(
                        ws, min_col=src_col, max_col=src_col,
                        min_row=hdr_row, max_row=table_end_row,
                    ),
                    titles_from_data=True,
                )
                ch.set_categories(
                    Reference(
                        ws, min_col=1, max_col=1,
                        min_row=hdr_row + 1, max_row=table_end_row,
                    )
                )
                _clean_axis(ch.x_axis)
                _clean_axis(ch.y_axis)
                ch.y_axis.numFmt = num_fmt
                ch.y_axis.scaling.min = 0
                max_v = max(series_vals) if series_vals else 0
                # max=0 durumunda 5 sabit; aksi halde %20 üst boşluk.
                ch.y_axis.scaling.max = (
                    float(max_v) * 1.2 if max_v > 0 else 5
                )
                ch.legend = None
                for s in ch.series:
                    s.marker = Marker(symbol="circle", size=5)
                    gp = GraphicalProperties()
                    gp.line = LineProperties(solidFill=line_color, w=22000)
                    s.graphicalProperties = gp
                ch.height = 8
                ch.width = 8
                _apply_chart_frame(ch)
                ws.add_chart(ch, f"{anchor}{hdr_row}")

            _mini("F", "Yarı Mamul Tonajı", 2, "#,##0", "EA580C", site_ton_vals)
            _mini("J", "Boş Konteyner", 3, "#,##0", "BE123C", site_empty_vals)
            _mini("N", "Dolu Konteyner", 4, "#,##0", "1F3A8A", site_full_vals)
            _mini("R", "Dolu Konteyner Tonajı", 5, "#,##0", "F59E0B", site_ton_vals)


# ---------------------------------------------------------------------------
# Üretim Yeri Karşılaştırma — weekly site-summary tables side by side
# ---------------------------------------------------------------------------

def _build_uretim_yeri_karsilastirma_sheet(
    wb: Workbook,
    all_rows: list[dict[str, Any]],
    manual_aggs: list[dict[str, Any]] | None = None,
) -> None:
    """Sheet 6: per-week ``Üretim Yeri Özeti`` tables stacked vertically.

    Each week renders its own table top-to-bottom (week-title row, header
    row, one row per production site, then a TOPLAM row). A two-row gap
    separates consecutive weeks. Newest week first.
    """
    ws = wb.create_sheet("Üretim Yeri Karşılaştırma")
    ws["A1"] = "Üretim Yeri Özeti — Haftalık Karşılaştırma"
    ws["A1"].font = Font(bold=True, size=14, color="1F3A8A")

    if not all_rows and not manual_aggs:
        ws["A3"] = "Henüz veri yok."
        return

    _, weekly_site, _, _, manual_only_weeks = _aggregate_all_weeks(
        all_rows, manual_aggs,
    )
    # Newest week at the top — admins typically open this sheet to compare
    # 'this week' against the recent past. Manual-only weeks (e.g. W17)
    # are included with a 'Manuel veri' tag in the title so admins know
    # the Kanban / Hurda / Tonaj columns will be zero by design (those
    # fields don't exist in the historical input).
    weeks = sorted(weekly_site.keys(), reverse=True)

    sub_headers = [
        "Üretim Yeri", "Boş", "Proseste", "Dolu", "Dolu içindeki Kanban",
        "Hurdaya ayrılacak",
        "Toplam Konteyner", "Toplam (%)", "Toplam Tonaj",
        "Dolu Konteyner Başına Yük",
    ]
    cols_per_table = len(sub_headers)
    gap_rows = 2

    # Column widths — set once for the whole sheet (same columns reused
    # by every weekly sub-table).
    ws.column_dimensions[get_column_letter(1)].width = 22
    for j in range(2, cols_per_table + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14

    start_row = 3
    for w in weeks:
        # Week title row, merged across the table width. Manual-only
        # weeks get a '(Manuel veri)' suffix so admins know the
        # Kanban / Hurda / Tonaj columns are zero by design.
        title_text = (
            f"{w} (Manuel veri)" if w in manual_only_weeks else w
        )
        title_cell = ws.cell(row=start_row, column=1, value=title_text)
        title_cell.font = Font(bold=True, size=12, color="1F3A8A")
        title_cell.alignment = _CENTER
        title_cell.fill = _TOTAL_FILL
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row, end_column=cols_per_table,
        )

        # Column header row — wrap text so long titles like 'Dolu
        # Konteyner Başına Yük' and 'Hurdaya ayrılacak' break onto a
        # second line instead of running into the next cell.
        header_row = start_row + 1
        ws.row_dimensions[header_row].height = 38
        wrap_center = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )
        for j, h in enumerate(sub_headers):
            cell = ws.cell(row=header_row, column=1 + j, value=h)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = wrap_center
            cell.border = _BORDER

        sites_in_week = weekly_site[w]
        grand_total_bdh = sum(
            s["empty"] + s.get("wip", 0) + s["full"] + s["scrap"]
            for s in sites_in_week.values()
        )
        totals = {"empty": 0, "wip": 0, "full": 0, "kanban": 0, "scrap": 0,
                  "bdh": 0, "tonnage": 0.0}

        for r_offset, (site, agg) in enumerate(
            sorted(
                sites_in_week.items(),
                key=lambda kv: _site_sort_key(kv[0]),
            ),
            start=header_row + 1,
        ):
            wip_v = agg.get("wip", 0)
            bdh = agg["empty"] + wip_v + agg["full"] + agg["scrap"]
            pct = (bdh / grand_total_bdh) if grand_total_bdh else 0
            ton_per = (agg["tonnage"] / agg["full"]) if agg["full"] else 0

            values = [
                site,
                agg["empty"], wip_v, agg["full"], agg["kanban"], agg["scrap"],
                bdh, pct, agg["tonnage"], ton_per,
            ]
            for j, val in enumerate(values):
                cell = ws.cell(row=r_offset, column=1 + j, value=val)
                cell.border = _BORDER
                if j == 0:
                    cell.alignment = _LEFT
                elif j == 7:  # Toplam (%)
                    cell.alignment = _RIGHT
                    cell.number_format = "0.0%"
                elif j == 9:  # Dolu Konteyner Başına Yük
                    cell.alignment = _RIGHT
                    cell.number_format = "0.00"
                else:
                    cell.alignment = _RIGHT
                    cell.number_format = "#,##0"

            totals["empty"] += agg["empty"]
            totals["wip"] += wip_v
            totals["full"] += agg["full"]
            totals["kanban"] += agg["kanban"]
            totals["scrap"] += agg["scrap"]
            totals["bdh"] += bdh
            totals["tonnage"] += agg["tonnage"]

        total_row = header_row + 1 + len(sites_in_week)
        ton_per_total = (totals["tonnage"] / totals["full"]) if totals["full"] else 0
        total_values = [
            "TOPLAM",
            totals["empty"], totals["wip"], totals["full"],
            totals["kanban"], totals["scrap"],
            totals["bdh"], 1.0, totals["tonnage"], ton_per_total,
        ]
        for j, val in enumerate(total_values):
            cell = ws.cell(row=total_row, column=1 + j, value=val)
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
            cell.border = _BORDER
            if j == 0:
                cell.alignment = _RIGHT
            elif j == 7:
                cell.alignment = _RIGHT
                cell.number_format = "0.0%"
            elif j == 9:
                cell.alignment = _RIGHT
                cell.number_format = "0.00"
            else:
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"

        # Move start_row past this week's table plus the gap.
        start_row = total_row + 1 + gap_rows

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Haftalık Analiz Özeti — geçen hafta vs bu hafta doğal dil özeti
# ---------------------------------------------------------------------------

_ANALIZ_METRICS = [
    # (key, human label, unit, num_fmt for delta values)
    ("full",    "Dolu Konteyner",             "adet", "#,##0"),
    ("empty",   "Boş Konteyner",              "adet", "#,##0"),
    ("wip",     "Proseste Konteyner",         "adet", "#,##0"),
    ("scrap",   "Hurdaya Ayrılacak Konteyner", "adet", "#,##0"),
    ("bdh",     "Toplam Konteyner",           "adet", "#,##0"),
    ("tonnage", "Toplam Tonaj",               "t",    "#,##0"),
]


def _build_haftalik_analiz_sheet(
    wb: Workbook,
    all_rows: list[dict[str, Any]],
    manual_aggs: list[dict[str, Any]] | None = None,
) -> None:
    """Executive summary — geçen hafta vs bu hafta karşılaştırması.

    Her metrik için:
      • Bir başlık satırı (renk şeridi)
      • Doğal dil özet cümle ('Geçen haftaya göre ... arttı/azaldı')
      • Üretim yeri kırılımı tablosu: Geçen Hafta / Bu Hafta / Fark / %
        (önce artanlar delta desc, sonra azalanlar delta asc)
    """
    ws = wb.create_sheet("Haftalık Analiz Özeti")

    # Banner
    ws.merge_cells("A1:F1")
    b = ws["A1"]
    b.value = "Haftalık Analiz Özeti"
    b.font = Font(bold=True, size=20, color="FFFFFF")
    b.fill = PatternFill("solid", fgColor="1F3A8A")
    b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 38

    _, weekly_site, _, _, manual_only_weeks = _aggregate_all_weeks(
        all_rows, manual_aggs,
    )
    weeks_full = sorted(
        w for w in weekly_site.keys() if w not in manual_only_weeks
    )
    if len(weeks_full) < 2:
        ws["A3"] = (
            "Karşılaştırma için en az iki haftalık veri gerekiyor — "
            "henüz yeterli veri yok."
        )
        ws["A3"].font = Font(italic=True, color="64748B")
        return

    latest_wk = weeks_full[-1]
    prev_wk = weeks_full[-2]

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = (
        f"Karşılaştırma: {_short_week(prev_wk)} → {_short_week(latest_wk)}"
    )
    sub.font = Font(italic=True, size=12, color="475569")
    sub.alignment = Alignment(
        horizontal="left", vertical="center", indent=1,
    )
    ws.row_dimensions[2].height = 22

    # Kolon genişlikleri — 6 kolon: Üretim Yeri | Önceki Ortalama |
    # Geçen Hafta | Bu Hafta | Fark | Değişim.
    ws.column_dimensions["A"].width = 26
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 15

    def _pct_txt(delta: float, base: float) -> str:
        if not base:
            return "—"
        p = delta / base * 100
        sign = "" if p == 0 else ("+" if p > 0 else "-")
        return f"{sign}%{_fmt_dec_tr(abs(p), 1)}"

    latest_sites = weekly_site.get(latest_wk, {})
    prev_sites = weekly_site.get(prev_wk, {})
    all_sites = sorted(
        set(latest_sites.keys()) | set(prev_sites.keys()),
        key=_site_sort_key,
    )

    def _fmt_delta_value(v: float, key: str) -> str:
        if key == "tonnage":
            return f"{_fmt_int_tr(abs(v))} t"
        return f"{_fmt_int_tr(abs(v))} adet"

    row = 4
    for key, label, unit, num_fmt in _ANALIZ_METRICS:
        # Totals
        prev_tot = sum(sd.get(key, 0) for sd in prev_sites.values())
        latest_tot = sum(sd.get(key, 0) for sd in latest_sites.values())
        delta = latest_tot - prev_tot
        pct_str = _pct_txt(delta, prev_tot)

        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        h = ws.cell(row=row, column=1, value=label)
        h.font = Font(bold=True, size=13, color="FFFFFF")
        h.fill = PatternFill("solid", fgColor="1F3A8A")
        h.alignment = Alignment(
            horizontal="left", vertical="center", indent=1,
        )
        ws.row_dimensions[row].height = 24
        row += 1

        # Summary sentence
        if delta > 0:
            verb = "arttı"
        elif delta < 0:
            verb = "azaldı"
        else:
            verb = "değişmedi"
        sentence = (
            f"Geçen haftaya göre toplam {label.lower()} "
            f"{_fmt_delta_value(delta, key)} olarak {verb} ({pct_str})."
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        sc = ws.cell(row=row, column=1, value=sentence)
        sc.font = Font(size=11, color="0F172A")
        sc.alignment = Alignment(
            horizontal="left", vertical="center", indent=1, wrap_text=True,
        )
        ws.row_dimensions[row].height = 26
        row += 2

        # Table header — 'Önceki Ortalama' bu haftadan önceki tüm
        # haftaların (prev_wk dahil) ortalamasını gösteriyor; kısa
        # vadeli oynama yerine geçmiş trend ile karşılaştırma sağlıyor.
        for j, ht in enumerate(
            ["Üretim Yeri", "Önceki Ortalama", "Geçen Hafta", "Bu Hafta",
             "Fark", "Değişim"],
            start=1,
        ):
            c = ws.cell(row=row, column=j, value=ht)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.alignment = _CENTER
            c.border = _BORDER

        # Bu haftaya kadar olan tüm önceki haftalar.
        past_weeks = weeks_full[:-1]

        site_deltas: list[tuple[str, float, float, float, float]] = []
        for site in all_sites:
            prv = float(prev_sites.get(site, {}).get(key, 0) or 0)
            lat = float(latest_sites.get(site, {}).get(key, 0) or 0)
            past_vals = [
                float(
                    weekly_site.get(w, {}).get(site, {}).get(key, 0) or 0
                )
                for w in past_weeks
            ]
            avg = (sum(past_vals) / len(past_vals)) if past_vals else 0.0
            d = lat - prv
            site_deltas.append((site, avg, prv, lat, d))

        increases = sorted(
            [x for x in site_deltas if x[4] > 0], key=lambda x: -x[4],
        )
        decreases = sorted(
            [x for x in site_deltas if x[4] < 0], key=lambda x: x[4],
        )
        unchanged = [x for x in site_deltas if x[4] == 0]

        r = row + 1
        for group in (increases, decreases, unchanged):
            for site, avg, prv, lat, d in group:
                zebra = _ZEBRA_FILL if (r - row) % 2 == 0 else None
                vals = [
                    site, avg, prv, lat, d, _pct_txt(d, prv),
                ]
                for j, v in enumerate(vals, start=1):
                    cell = ws.cell(row=r, column=j, value=v)
                    cell.border = _BORDER
                    if zebra:
                        cell.fill = zebra
                    if j == 1:
                        cell.alignment = _LEFT
                    elif j == 6:  # Değişim yüzde string
                        cell.alignment = _RIGHT
                        # Delta > 0 → yeşil, < 0 → kırmızı, = 0 → gri
                        if d > 0:
                            cell.font = Font(bold=True, color="047857")
                        elif d < 0:
                            cell.font = Font(bold=True, color="BE123C")
                        else:
                            cell.font = Font(color="64748B")
                    else:
                        cell.alignment = _RIGHT
                        cell.number_format = num_fmt
                        if j == 5 and d != 0:  # Fark sütunu — renkli
                            cell.font = Font(
                                bold=True,
                                color="047857" if d > 0 else "BE123C",
                            )
                r += 1

        row = r + 2  # section gap


# ---------------------------------------------------------------------------
# Ana Data Sayfası — tüm haftaların uzun-format kaydı, pivot için
# ---------------------------------------------------------------------------

def _build_ana_data_sheet(
    wb: Workbook, all_weeks_rows: list[dict[str, Any]],
) -> None:
    """Long-format ham veri sayfası — her (hafta × üretim yeri × bölüm ×
    renk) bir satır. Kullanıcı bu sayfayı pivot table kaynağı olarak
    kullanıp analiz sayfalarında olmayan soruları cevaplayabiliyor."""
    ws = wb.create_sheet("Ana Data Sayfası")
    headers = [
        "Hafta", "Hafta Aralığı", "Ay", "Yıl",
        "Üretim Yeri", "Bölüm", "Renk",
        "Boş", "Proseste", "Dolu", "Kanban", "Hurdaya Ayrılacak",
        "Toplam Konteyner",
        "Durum", "Giren Kullanıcı",
        "Sayım Tarihi", "Sayım Saati", "Gönderim Zamanı",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    if not all_weeks_rows:
        return

    for idx, row in enumerate(all_weeks_rows, start=2):
        is_late = row.get("Durum") == "late_submitted"
        zebra = (idx % 2 == 0) and not is_late
        fill = _LATE_FILL if is_late else (_ZEBRA_FILL if zebra else None)

        week_iso = row.get("Hafta") or ""
        hafta_araligi, ay, yil = _week_iso_to_human(week_iso)

        bos_v = int(row.get("Boş") or 0)
        wip_v = int(row.get("Proseste") or 0)
        dolu_v = int(row.get("Dolu") or 0)
        hurda_v = int(row.get("Hurda") or 0)
        bdh_v = bos_v + wip_v + dolu_v + hurda_v

        values = [
            week_iso, hafta_araligi, ay, yil if yil else "",
            row.get("Üretim Yeri", ""), row.get("Bölüm", ""),
            row.get("Renk", ""),
            row.get("Boş"), row.get("Proseste"),
            row.get("Dolu"), row.get("Kanban"), row.get("Hurda"),
            bdh_v,
            _STATUS_LABEL.get(row.get("Durum"), row.get("Durum", "")),
            row.get("Giren Kullanıcı", ""),
            row.get("Sayım Tarihi", ""), row.get("Sayım Saati", ""),
            _fmt_ts(row.get("Gönderim Zamanı")),
        ]
        ws.append(values)

        for col_idx in range(1, len(values) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.border = _BORDER
            if fill:
                cell.fill = fill
            # 8=Boş, 9=WIP, 10=Dolu, 11=Kanban, 12=Hurda, 13=Toplam
            if col_idx in (8, 9, 10, 11, 12, 13):
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            elif col_idx == 14:  # Durum
                cell.alignment = _CENTER
                if is_late:
                    cell.font = Font(bold=True, color="92400E")
            elif col_idx in (1, 4):
                cell.alignment = _CENTER
            else:
                cell.alignment = _LEFT

    _autofit(ws, headers)


# ---------------------------------------------------------------------------
# Dolu Konteyner Başına Yük Özeti — sites × weeks matrix
# ---------------------------------------------------------------------------

def _build_dolu_yuk_ozeti_sheet(
    wb: Workbook,
    all_rows: list[dict[str, Any]],
    manual_aggs: list[dict[str, Any]] | None = None,
) -> None:
    """Per-site weekly ton/Dolu Konteyner; weeks shown side by side.

    Satırlar üretim yerleri, sütunlar haftalar (yeniden eskiye), her
    hücre o tesisin o haftaki ``tonaj / dolu konteyner`` oranını gösterir.
    Son sütun tesis için tüm-hafta ağırlıklı ortalaması; alt satır
    haftalık ağırlıklı ortalamaları + genel ortalamayı verir.
    """
    ws = wb.create_sheet("Dolu Konteyner Başına Yük Özeti")

    if not all_rows and not manual_aggs:
        ws["A1"] = "Henüz veri yok."
        ws["A1"].font = Font(italic=True, color="64748B")
        return

    _, weekly_site, _, _, manual_only_weeks = _aggregate_all_weeks(
        all_rows, manual_aggs,
    )
    # Sadece tonajı olan haftalar — manual-only haftalarda tonaj/dolu
    # yok, sütun olarak gösterirsek hep boş kalır. Kronolojik sıra:
    # en eski hafta solda, son hafta sağda — sola doğru okurken
    # zamanda ileri gidiyoruz.
    weeks = sorted(
        w for w in weekly_site.keys() if w not in manual_only_weeks
    )
    all_sites = sorted(
        {s for sd in weekly_site.values() for s in sd.keys()},
        key=_site_sort_key,
    )

    if not weeks or not all_sites:
        ws["A1"] = "Henüz tonajlı sayım verisi yok."
        ws["A1"].font = Font(italic=True, color="64748B")
        return

    headers = ["Üretim Yeri"] + [_short_week(w) for w in weeks] + ["Ortalama"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for idx, site in enumerate(all_sites, start=2):
        row_vals: list[Any] = [site]
        sum_ton, sum_full = 0.0, 0
        for w in weeks:
            sd = weekly_site.get(w, {}).get(site)
            full_v = int(sd.get("full", 0)) if sd else 0
            ton_v = float(sd.get("tonnage", 0.0)) if sd else 0.0
            if full_v:
                row_vals.append(ton_v / full_v)
                sum_ton += ton_v
                sum_full += full_v
            else:
                row_vals.append(None)
        row_vals.append((sum_ton / sum_full) if sum_full else None)
        ws.append(row_vals)

        zebra = _ZEBRA_FILL if idx % 2 == 0 else None
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.border = _BORDER
            if zebra:
                cell.fill = zebra
            if col_idx == 1:
                cell.alignment = _LEFT
            else:
                cell.alignment = _RIGHT
                cell.number_format = "0.00"

    # TOPLAM satırı — haftalık ve genel ağırlıklı ortalama.
    total_row_idx = ws.max_row + 1
    total_vals: list[Any] = ["TOPLAM"]
    all_ton, all_full = 0.0, 0
    for w in weeks:
        week_ton, week_full = 0.0, 0
        for site in all_sites:
            sd = weekly_site.get(w, {}).get(site)
            if sd:
                week_ton += float(sd.get("tonnage", 0.0))
                week_full += int(sd.get("full", 0))
        total_vals.append((week_ton / week_full) if week_full else None)
        all_ton += week_ton
        all_full += week_full
    total_vals.append((all_ton / all_full) if all_full else None)
    ws.append(total_vals)

    for col_idx in range(1, len(total_vals) + 1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _BORDER
        if col_idx == 1:
            cell.alignment = _RIGHT
        else:
            cell.alignment = _RIGHT
            cell.number_format = "0.00"

    # freeze_panes kaldırıldı (hem 'B2' hem _style_header_row'un
    # verdiği 'A2') — altta çizilen grafiklerin görüntü alanını
    # bozuyordu.
    ws.freeze_panes = None
    _autofit(ws, headers)
    ws.column_dimensions["A"].width = 22
    n_cols = 1 + len(weeks) + 1
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8

    # ================================================================
    # Grafik alanı — tablonun altında haftalık ham tonaj görselleştirmesi.
    # Ana grafik: haftalık toplam ham tonaj. Yanında üretim yeri
    # butonları; tıklanınca ilgili tesisin per-tesis bloğuna atlıyor.
    # Per-tesis blokları ana grafiğin altında dikey stack.
    # ================================================================
    n_weeks = len(weeks)
    n_sites = len(all_sites)

    # Bölüm başlığı
    sec_row = total_row_idx + 3
    ws.merge_cells(
        start_row=sec_row, start_column=1,
        end_row=sec_row, end_column=n_cols,
    )
    sec = ws.cell(
        row=sec_row, column=1,
        value="Haftalık Yarı Mamul Tonajı Görselleştirmesi",
    )
    sec.font = Font(bold=True, size=13, color="1F3A8A")
    sec.fill = PatternFill("solid", fgColor="E2E8F0")
    sec.alignment = Alignment(
        horizontal="left", vertical="center", indent=1,
    )
    ws.row_dimensions[sec_row].height = 24

    # Gizli veri alanı — chart Reference'ları için sağda kolonlar.
    hidden_col_weeks = 40
    hidden_col_total = 41
    ws.cell(row=1, column=hidden_col_weeks, value="Hafta")
    ws.cell(row=1, column=hidden_col_total, value="Toplam Tonaj")
    for i, s in enumerate(all_sites):
        ws.cell(row=1, column=42 + i, value=s)

    for k, w in enumerate(weeks):
        ws.cell(row=2 + k, column=hidden_col_weeks, value=_short_week(w))
        total_ton = sum(
            float(
                weekly_site.get(w, {}).get(s, {}).get("tonnage", 0.0) or 0
            )
            for s in all_sites
        )
        c_tot = ws.cell(
            row=2 + k, column=hidden_col_total, value=total_ton,
        )
        c_tot.number_format = "#,##0"
        for i, s in enumerate(all_sites):
            val = float(
                weekly_site.get(w, {}).get(s, {}).get("tonnage", 0.0) or 0
            )
            cell = ws.cell(row=2 + k, column=42 + i, value=val)
            cell.number_format = "#,##0"
    hidden_last_row = 1 + n_weeks

    main_chart_anchor = sec_row + 2
    per_site_block_rows = 20  # banner(1) + back(1) + chart ~16 + gap 2
    per_site_start_row = main_chart_anchor + 22
    site_anchors_local: dict[str, int] = {
        s: per_site_start_row + i * per_site_block_rows
        for i, s in enumerate(all_sites)
    }

    # Ana grafik — mavi BarChart, gapWidth=30, y-axis title yok.
    main_chart = BarChart()
    main_chart.type = "col"
    main_chart.style = 2
    main_chart.title = _make_chart_title(
        "Haftalık Toplam Yarı Mamul Tonajı (Tüm Tesisler)"
    )
    main_chart.x_axis.title = _end_x_axis_title("Hafta")
    if weeks:
        main_chart.add_data(
            Reference(
                ws, min_col=hidden_col_total, min_row=1,
                max_col=hidden_col_total, max_row=hidden_last_row,
            ),
            titles_from_data=True,
        )
        main_chart.set_categories(
            Reference(
                ws, min_col=hidden_col_weeks, min_row=2,
                max_row=hidden_last_row,
            )
        )
    _clean_axis(main_chart.x_axis)
    _clean_axis(main_chart.y_axis)
    main_chart.y_axis.numFmt = "#,##0"
    main_chart.y_axis.scaling.min = 0
    main_chart.gapWidth = 30
    for series in main_chart.series:
        gp = GraphicalProperties(solidFill="1F3A8A")
        series.graphicalProperties = gp
    main_chart.dataLabels = _value_only_labels(
        "t", "#,##0",
        txPr=_bold_large_label_props(size_pt=10, color="0F172A"),
    )
    main_chart.legend = None
    main_chart.height = 10
    main_chart.width = 22
    _apply_chart_frame(main_chart)
    ws.add_chart(main_chart, f"A{main_chart_anchor}")

    # Yandaki üretim yeri butonları — target_cell doğrudan blok
    # banner satırına (offset yok). Excel scroll edince banner
    # satırı visible area'nın en üstünde çıkıyor; kullanıcı ilk
    # olarak site adını görüyor, üstünde önceki bloğun chart
    # kalıntısı olmuyor.
    btn_col = 16
    btn_h = 2
    chart_span_rows = 20
    total_btn_h = n_sites * btn_h
    start_offset = max(0, (chart_span_rows - total_btn_h) // 2)
    for i, s in enumerate(all_sites):
        _link_button_excel(
            ws,
            row=main_chart_anchor + start_offset + i * btn_h,
            col=btn_col, width=4, height=btn_h,
            label=s,
            target_sheet="Dolu Konteyner Başına Yük Özeti",
            target_cell=f"A{site_anchors_local[s]}",
            font_size=10,
        )

    # Per-tesis blokları — banner + back link + BarChart.
    for i, s in enumerate(all_sites):
        blk = site_anchors_local[s]

        ws.merge_cells(
            start_row=blk, start_column=1, end_row=blk, end_column=n_cols,
        )
        b = ws.cell(
            row=blk, column=1,
            value=f"{s} — Haftalık Yarı Mamul Tonajı",
        )
        b.font = Font(bold=True, size=12, color="FFFFFF")
        b.fill = PatternFill("solid", fgColor="1F3A8A")
        b.alignment = Alignment(
            horizontal="left", vertical="center", indent=1,
        )
        ws.row_dimensions[blk].height = 24

        bk = ws.cell(
            row=blk + 1, column=1, value="◀ Ana grafiğe dön",
        )
        # Back link doğrudan ana grafik anchor'ına (offset yok) —
        # kullanıcı geri dönünce hemen ana grafiği görüyor.
        bk.hyperlink = Hyperlink(
            ref=bk.coordinate,
            location=(
                f"'Dolu Konteyner Başına Yük Özeti'!A{main_chart_anchor}"
            ),
            display="◀ Ana grafiğe dön",
        )
        bk.font = Font(
            bold=True, color="1F3A8A", size=10, underline="single",
        )
        bk.alignment = Alignment(horizontal="left", vertical="center")

        # Per-tesis mavi BarChart
        site_col = 42 + i
        ch = BarChart()
        ch.type = "col"
        ch.style = 2
        ch.title = _make_chart_title(f"{s} — Haftalık Yarı Mamul Tonajı")
        ch.x_axis.title = _end_x_axis_title("Hafta")
        if weeks:
            ch.add_data(
                Reference(
                    ws, min_col=site_col, min_row=1,
                    max_col=site_col, max_row=hidden_last_row,
                ),
                titles_from_data=True,
            )
            ch.set_categories(
                Reference(
                    ws, min_col=hidden_col_weeks, min_row=2,
                    max_row=hidden_last_row,
                )
            )
        _clean_axis(ch.x_axis)
        _clean_axis(ch.y_axis)
        ch.y_axis.numFmt = "#,##0"
        ch.y_axis.scaling.min = 0
        # Dar bant değerlerde tick label tekrarını engellemek için
        # scaling.max manuel.
        site_vals = [
            float(
                weekly_site.get(w, {}).get(s, {}).get("tonnage", 0.0) or 0
            )
            for w in weeks
        ]
        max_v = max(site_vals) if site_vals else 0
        ch.y_axis.scaling.max = max_v * 1.2 if max_v > 0 else 5
        ch.gapWidth = 30
        for series in ch.series:
            gp = GraphicalProperties(solidFill="1F3A8A")
            series.graphicalProperties = gp
        ch.dataLabels = _value_only_labels(
            "t", "#,##0",
            txPr=_bold_large_label_props(size_pt=9, color="0F172A"),
        )
        ch.legend = None
        ch.height = 8
        ch.width = 22
        _apply_chart_frame(ch)
        ws.add_chart(ch, f"A{blk + 2}")


def _build_yari_mamul_tonaj_ozeti_sheet(
    wb: Workbook,
    all_rows: list[dict[str, Any]],
    manual_aggs: list[dict[str, Any]] | None = None,
) -> None:
    """Per-site weekly yarı mamul (ham) tonaj matrisi.

    Dolu Yük Özeti ile aynı yapı ama hücreler ``ton/dolu`` oranı
    değil ham gerçekleşen tonaj toplamı. Satırlar üretim yerleri,
    sütunlar haftalar kronolojik. Son sütun tesis toplamı; alt
    satır haftalık toplam + genel toplam.
    """
    ws = wb.create_sheet("Yarı Mamul Tonajı Özeti")

    if not all_rows and not manual_aggs:
        ws["A1"] = "Henüz veri yok."
        ws["A1"].font = Font(italic=True, color="64748B")
        return

    _, weekly_site, _, _, manual_only_weeks = _aggregate_all_weeks(
        all_rows, manual_aggs,
    )
    weeks = sorted(
        w for w in weekly_site.keys() if w not in manual_only_weeks
    )
    all_sites = sorted(
        {s for sd in weekly_site.values() for s in sd.keys()},
        key=_site_sort_key,
    )

    if not weeks or not all_sites:
        ws["A1"] = "Henüz tonajlı sayım verisi yok."
        ws["A1"].font = Font(italic=True, color="64748B")
        return

    # Son sütun: son haftanın önceki 3 haftanın ortalamasına göre
    # farkı. Örn. W28 sayımında W25/W26/W27 ortalamasıyla
    # karşılaştırılıyor. Yeterli hafta yoksa (< 2) sütunu atlıyoruz.
    latest_wk = weeks[-1]
    prev_weeks = weeks[-4:-1] if len(weeks) >= 4 else weeks[:-1]
    show_delta_col = bool(prev_weeks)

    if show_delta_col:
        delta_label = (
            f"{_short_week(latest_wk)} − "
            f"{_short_week(prev_weeks[0])}…{_short_week(prev_weeks[-1])} Ort."
        )
        headers = (
            ["Üretim Yeri"]
            + [_short_week(w) for w in weeks]
            + [delta_label]
        )
    else:
        headers = ["Üretim Yeri"] + [_short_week(w) for w in weeks]
    ws.append(headers)
    _style_header_row(
        ws, len(headers),
        wrap_text=True, row_height=(42 if show_delta_col else 26),
    )

    for idx, site in enumerate(all_sites, start=2):
        row_vals: list[Any] = [site]
        latest_ton = 0.0
        prev_tons: list[float] = []
        for w in weeks:
            sd = weekly_site.get(w, {}).get(site)
            ton_v = float(sd.get("tonnage", 0.0)) if sd else 0.0
            row_vals.append(ton_v if ton_v else None)
            if w == latest_wk:
                latest_ton = ton_v
            if w in prev_weeks:
                prev_tons.append(ton_v)

        delta_val: float | None = None
        if show_delta_col and prev_tons:
            prev_avg = sum(prev_tons) / len(prev_tons)
            delta_val = latest_ton - prev_avg
            row_vals.append(delta_val)

        ws.append(row_vals)

        zebra = _ZEBRA_FILL if idx % 2 == 0 else None
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.border = _BORDER
            if zebra:
                cell.fill = zebra
            if col_idx == 1:
                cell.alignment = _LEFT
            else:
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
                # Delta sütunu: pozitif yeşil, negatif kırmızı,
                # imzalı format (Excel'in beğendiği quote'lu form).
                if show_delta_col and col_idx == len(row_vals) \
                        and delta_val is not None and delta_val != 0:
                    cell.number_format = "\"+\"#,##0;\"-\"#,##0"
                    cell.font = Font(
                        bold=True,
                        color="047857" if delta_val > 0 else "BE123C",
                    )

    # TOPLAM satırı — haftalık toplam + toplu delta.
    total_row_idx = ws.max_row + 1
    total_vals: list[Any] = ["TOPLAM"]
    latest_total = 0.0
    prev_totals: list[float] = []
    for w in weeks:
        week_total = sum(
            float(weekly_site.get(w, {}).get(s, {}).get("tonnage", 0.0) or 0)
            for s in all_sites
        )
        total_vals.append(week_total if week_total else None)
        if w == latest_wk:
            latest_total = week_total
        if w in prev_weeks:
            prev_totals.append(week_total)

    tot_delta: float | None = None
    if show_delta_col and prev_totals:
        tot_delta = latest_total - sum(prev_totals) / len(prev_totals)
        total_vals.append(tot_delta)

    ws.append(total_vals)

    for col_idx in range(1, len(total_vals) + 1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _BORDER
        if col_idx == 1:
            cell.alignment = _RIGHT
        else:
            cell.alignment = _RIGHT
            cell.number_format = "#,##0"
            if show_delta_col and col_idx == len(total_vals) \
                    and tot_delta is not None and tot_delta != 0:
                cell.number_format = "\"+\"#,##0;\"-\"#,##0"
                cell.font = Font(
                    bold=True,
                    color="047857" if tot_delta > 0 else "BE123C",
                )

    # freeze_panes kaldırıldı — ilk sütun dondurulunca yatay scroll
    # da farklı davranıyor, kullanıcı istediği için None.
    ws.freeze_panes = None
    _autofit(ws, headers)
    ws.column_dimensions["A"].width = 22
    n_cols = len(headers)
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9
    # Delta sütunu daha geniş — başlığı uzun.
    if show_delta_col:
        ws.column_dimensions[get_column_letter(n_cols)].width = 14


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_week_excel(
    rows: list[dict[str, Any]],
    week_iso: str,
    week_human: str,
    all_weeks_rows: list[dict[str, Any]] | None = None,
    manual_aggs: list[dict[str, Any]] | None = None,
) -> bytes:
    """Return an .xlsx byte string for the selected week.

    ``rows``           — output of ``get_week_export_rows(week_iso)``.
    ``all_weeks_rows`` — output of ``get_all_weeks_export_rows()``; used
                        for the GRAFİKLER (charts) and Üretim Yeri
                        Karşılaştırma sheets.
    ``manual_aggs``    — output of ``get_manual_site_aggregates()``;
                        per-(week, site) totals for historical weeks
                        counted outside the normal flow. Folded into
                        the charts + comparison sheet via
                        ``_aggregate_all_weeks``.
    ``week_iso`` and ``week_human`` are accepted for backwards-compatible
    call sites but no longer surfaced in a 'Bilgi' cover sheet.
    """
    wb = Workbook()
    # Workbook() default boş 'Sheet' yaratıyor; artık Dashboard sayfası
    # yok, ilk gerçek sayfa Renk Kırılımı olacak — boş Sheet'i sil.
    default_ws = wb.active
    wb.remove(default_ws)

    # Executive summary — hafta karşılaştırması + doğal dil özet cümleler.
    # 0. sırada oluşturuluyor ki dosya açılınca ilk gelen sayfa bu olsun.
    _build_haftalik_analiz_sheet(
        wb, all_weeks_rows or [], manual_aggs or [],
    )
    _build_renk_kirilim_sheet(wb, rows)
    dept_aggs = _build_uretim_yeri_kirilim_sheet(wb, rows)
    _build_uretim_yeri_ozeti_sheet(wb, dept_aggs)
    _build_renk_ozeti_sheet(wb, rows)
    _build_dolu_yuk_ozeti_sheet(wb, all_weeks_rows or [], manual_aggs or [])
    _build_yari_mamul_tonaj_ozeti_sheet(
        wb, all_weeks_rows or [], manual_aggs or [],
    )
    _build_uretim_yeri_karsilastirma_sheet(
        wb, all_weeks_rows or [], manual_aggs or []
    )
    # Pivot için ham veri — kullanıcı analiz sayfalarında olmayan
    # soruları buradan cevaplayabilir.
    _build_ana_data_sheet(wb, all_weeks_rows or [])
    # Grafikler en son sheet olarak kalsın — üretim yeri trend
    # bloklarını kendi içinde, butonlarla erişilen 'Tesis Detayı'
    # bölümünde tutuyor.
    _build_ozet_charts_sheet(wb, all_weeks_rows or [], manual_aggs or [])

    # Analiz sayfasını 0. sıraya taşı (create_sheet ordinal 0 vermek
    # her zaman çalışmıyor).
    if "Haftalık Analiz Özeti" in wb.sheetnames:
        idx = wb.sheetnames.index("Haftalık Analiz Özeti")
        if idx != 0:
            wb.move_sheet("Haftalık Analiz Özeti", offset=-idx)
        wb.active = 0

    # Klavuz çizgilerini kapat — workbook bittiğinde dosya bir rapor
    # gibi görünsün, ham tablo gibi değil. (View > Gridlines'ın isteğe
    # bağlı işaretlenmesinin tersi.)
    for _ws in wb.worksheets:
        _ws.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_all_weeks_excel(rows: list[dict[str, Any]]) -> bytes:
    """Long-format export — one row per (week × site × department × color).

    Mirrors ``build_week_excel`` styling so it reads consistently with
    the per-week file, but adds Hafta / Hafta Aralığı / Ay / Yıl columns
    upfront so the user can pivot or filter quickly in Excel.
    """
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Tüm Haftalar"

    headers = [
        "Hafta",
        "Hafta Aralığı",
        "Ay",
        "Yıl",
        "Üretim Yeri",
        "Bölüm",
        "Renk",
        "Boş",
        "Proseste",
        "Dolu",
        "Kanban",
        "Hurdaya Ayrılacak",
        "Toplam Konteyner",
        "Gerçekleşen Tonaj (t)",
        "Durum",
        "Giren Kullanıcı",
        "Sayım Tarihi",
        "Sayım Saati",
        "Gönderim Zamanı",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    sheet.row_dimensions[1].height = 26
    sheet.freeze_panes = "A2"

    for idx, row in enumerate(rows, start=2):
        is_late = row.get("Durum") == "late_submitted"
        zebra = (idx % 2 == 0) and not is_late
        fill = _LATE_FILL if is_late else (_ZEBRA_FILL if zebra else None)

        week_iso = row.get("Hafta") or ""
        hafta_araligi, ay, yil = _week_iso_to_human(week_iso)

        bos_v = int(row.get("Boş") or 0)
        wip_v = int(row.get("Proseste") or 0)
        dolu_v = int(row.get("Dolu") or 0)
        hurda_v = int(row.get("Hurda") or 0)
        bdh_v = bos_v + wip_v + dolu_v + hurda_v

        values = [
            week_iso,
            hafta_araligi,
            ay,
            yil if yil else "",
            row.get("Üretim Yeri", ""),
            row.get("Bölüm", ""),
            row.get("Renk", ""),
            row.get("Boş"),
            row.get("Proseste"),
            row.get("Dolu"),
            row.get("Kanban"),
            row.get("Hurda"),
            bdh_v,
            row.get("Gerçekleşen Tonaj"),
            _STATUS_LABEL.get(row.get("Durum"), row.get("Durum", "")),
            row.get("Giren Kullanıcı", ""),
            row.get("Sayım Tarihi", ""),
            row.get("Sayım Saati", ""),
            _fmt_ts(row.get("Gönderim Zamanı")),
        ]
        sheet.append(values)

        for col_idx in range(1, len(values) + 1):
            cell = sheet.cell(row=idx, column=col_idx)
            cell.border = _BORDER
            if fill:
                cell.fill = fill
            # Sayısal: 8=Boş, 9=WIP, 10=Dolu, 11=Kanban, 12=Hurda, 13=Toplam, 14=Tonaj
            if col_idx in (8, 9, 10, 11, 12, 13, 14):
                cell.alignment = _RIGHT
                cell.number_format = "#,##0"
            elif col_idx == 15:  # Durum
                cell.alignment = _CENTER
                if is_late:
                    cell.font = Font(bold=True, color="92400E")
            elif col_idx in (1, 4):
                cell.alignment = _CENTER
            else:
                cell.alignment = _LEFT

    _autofit(sheet, headers)

    # Cover sheet
    info = wb.create_sheet("Bilgi", 0)
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 60
    info["A1"] = "Norm Konteyner — Tüm Haftalar Sayım Raporu"
    info["A1"].font = Font(bold=True, size=16, color="1F3A8A")
    info.merge_cells("A1:B1")
    info["A3"] = "Toplam satır"
    info["B3"] = len(rows)
    weeks = sorted({r.get("Hafta") for r in rows if r.get("Hafta")}, reverse=True)
    info["A4"] = "Hafta sayısı"
    info["B4"] = len(weeks)
    if weeks:
        info["A5"] = "En yeni hafta"
        info["B5"] = weeks[0]
        info["A6"] = "En eski hafta"
        info["B6"] = weeks[-1]
    info["A7"] = "Oluşturulma"
    info["B7"] = now_tr().strftime("%Y-%m-%d %H:%M")
    for r in range(3, 8):
        info.cell(row=r, column=1).font = Font(bold=True)
        info.cell(row=r, column=1).alignment = _LEFT
        info.cell(row=r, column=2).alignment = _LEFT

    # Klavuz çizgilerini kapat — kalan workbook gibi temiz dursun.
    for _ws in wb.worksheets:
        _ws.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
